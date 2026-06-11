"""
Ta'awon P3 SMS - FastAPI Backend
================================
Single-file backend.

Changes in v0.8 (3 May 2026, "real authentication"):
  + Session-based authentication: login issues a session cookie that contains
    a random opaque token; the server stores active sessions in a new table
    `Sessions` and looks them up on every protected request.
  + Two parallel auth tracks:
      * App_Users (internal staff)        -> POST /api/v1/auth/login
      * External_Submitters (external)    -> POST /api/v1/auth/external-login
    Both produce the same kind of session cookie ("sms_session"); the server
    distinguishes them via a `principal_type` field on the session row.
  + bcrypt password verification for both tracks. The seeded staff users now
    have their `temp_hash_123` placeholder hash auto-migrated to a real bcrypt
    hash of the literal string "ChangeMe123!" on init-db, with
    force_password_change=TRUE so they MUST change it on first login.
  + Force password change: if `force_password_change` (staff) or
    `must_change_password` (external) is TRUE, the server allows ONLY the
    /api/v1/auth/change-password endpoint until the user actually changes it.
  + Sessions last 8 hours from issue; refreshed on every request; revoked on
    logout. A simple cleanup query removes expired rows on each login.
  + Removed reliance on X-Acting-User-Id header. Old code paths that used it
    are gone; the server now derives identity from the session cookie alone.
  + Admin-create-staff-user now generates a random temporary password and
    returns it ONCE in the response body (admin copies and shares manually).
    The seeded `temp_hash_123` placeholder for new users is gone.
  + External-submitter creation also returns the temp password in the
    response (since email delivery is unreliable on the current host).
  + Auth-gated attachment downloads: only authenticated users with
    view_all_applications can download, plus the original submitter of an
    application can download their own attachments.
  + Rate-limit on login: 5 failed attempts in 15 minutes per (username, IP)
    triggers a 15-minute cooldown.
  + Restricted view for external submitters enforced server-side: they can
    ONLY hit /auth/me, /auth/change-password, /auth/logout, /organizations
    (their own only), and /applications/submit. Everything else returns 403.

Migrations:
  - Sessions table created.
  - failed_login_attempts table created.
  - App_Users.hashed_password updated for any row still equal to
    'temp_hash_123' (one-time, idempotent).
  - The X-Acting-User-Id-based admin endpoints continue to exist in name but
    now read identity from the session cookie instead.
"""

import os
import secrets
import smtplib
import ssl
import sys
import uuid
from datetime import datetime, date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO
from pathlib import Path
from typing import List, Optional, Tuple

import asyncpg
import bcrypt
from dotenv import load_dotenv
from fastapi import (Cookie, Depends, FastAPI, File, Form, HTTPException, Path as FPath,
                     Request, Response, UploadFile)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

# ============================================================
#  Configuration
# ============================================================
PROJECT_ROOT = Path(__file__).resolve().parent
ENV_FILE = PROJECT_ROOT / ".env"
load_dotenv(ENV_FILE)
load_dotenv(Path(__file__).resolve().parent / ".env")

DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
DB_HOST     = os.getenv("DB_HOST", "localhost")
DB_PORT     = int(os.getenv("DB_PORT", "5432"))
DB_NAME     = os.getenv("DB_NAME", "id_system")
DB_USER     = os.getenv("DB_USER", "postgres")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")

if not DATABASE_URL and not DB_PASSWORD:
    print("ERROR: No database configuration found.", file=sys.stderr)
    print("Set either DATABASE_URL (preferred for production) or DB_PASSWORD.", file=sys.stderr)
    sys.exit(1)

SMTP_HOST     = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT     = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
SMTP_FROM     = os.getenv("SMTP_FROM", SMTP_USERNAME)
SMTP_ENABLED  = bool(SMTP_USERNAME and SMTP_PASSWORD)

_cors_raw = os.getenv("CORS_ORIGINS", "*").strip()
CORS_ORIGINS = [o.strip() for o in _cors_raw.split(",") if o.strip()] or ["*"]

UPLOADS_DIR = Path(os.getenv("UPLOADS_DIR", str(PROJECT_ROOT / "uploads")))
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
ALLOWED_MIME_PREFIXES = ("image/",)
ALLOWED_MIME_EXACT = {"application/pdf"}

# MinIO / S3-compatible storage (Railway MinIO add-on)
# Accepts both Railway MinIO add-on names and standard MinIO names
MINIO_ENDPOINT   = os.getenv("MINIO_ENDPOINT", os.getenv("AWS_ENDPOINT_URL", ""))
MINIO_ACCESS_KEY = os.getenv("MINIO_ACCESS_KEY") or os.getenv("MINIO_ROOT_USER", os.getenv("AWS_ACCESS_KEY_ID", ""))
MINIO_SECRET_KEY = os.getenv("MINIO_SECRET_KEY") or os.getenv("MINIO_ROOT_PASSWORD", os.getenv("AWS_SECRET_ACCESS_KEY", ""))
MINIO_BUCKET     = os.getenv("MINIO_BUCKET", os.getenv("AWS_S3_BUCKET", "attachments"))
USE_S3 = bool(MINIO_ENDPOINT)

if USE_S3:
    import boto3
    try:
        endpoint = MINIO_ENDPOINT if MINIO_ENDPOINT.startswith("http") else f"http://{MINIO_ENDPOINT}"
        s3_client = boto3.client(
            "s3",
            endpoint_url=endpoint,
            aws_access_key_id=MINIO_ACCESS_KEY,
            aws_secret_access_key=MINIO_SECRET_KEY,
            region_name=os.getenv("AWS_REGION", "us-east-1"),
        )
        s3_client.create_bucket(Bucket=MINIO_BUCKET)
        print(f"MinIO/S3 bucket '{MINIO_BUCKET}' ready at {endpoint}")
    except Exception as exc:
        print(f"WARNING: MinIO/S3 init failed ({exc}). Falling back to local filesystem.", file=sys.stderr)
        s3_client = None
        USE_S3 = False
else:
    s3_client = None

# Session config
SESSION_COOKIE_NAME = "sms_session"
SESSION_DURATION_HOURS = 8
SESSION_DURATION = timedelta(hours=SESSION_DURATION_HOURS)
LOGIN_THROTTLE_WINDOW = timedelta(minutes=15)
LOGIN_THROTTLE_MAX_ATTEMPTS = 5

# Default seeded password for staff on first deployment.
# Forced to be changed on first login.
DEFAULT_SEED_PASSWORD = "ChangeMe123!"

# ============================================================
#  App setup
# ============================================================
app = FastAPI(title="Ta'awon P3 SMS API", version="0.9.0")

# Mount SpeakEasy static app at /speakeasy
SPEAKEASY_DIR = PROJECT_ROOT / "speakeasy"
if SPEAKEASY_DIR.is_dir():
    app.mount("/speakeasy", StaticFiles(directory=str(SPEAKEASY_DIR), html=True), name="speakeasy")

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,   # required for cookies to round-trip
    allow_methods=["*"],
    allow_headers=["*"],
)


async def get_conn() -> asyncpg.Connection:
    if DATABASE_URL:
        return await asyncpg.connect(DATABASE_URL)
    return await asyncpg.connect(
        host=DB_HOST, port=DB_PORT, database=DB_NAME,
        user=DB_USER, password=DB_PASSWORD
    )


# ============================================================
#  Schema bootstrap (additive - safe to re-run)
# ============================================================
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS App_Users (
    user_id              SERIAL PRIMARY KEY,
    full_name            VARCHAR(100) NOT NULL,
    username             VARCHAR(50) UNIQUE NOT NULL,
    hashed_password      VARCHAR(255) NOT NULL,
    force_password_change BOOLEAN DEFAULT TRUE,
    role_id              INT DEFAULT 1,
    is_active            BOOLEAN DEFAULT TRUE,
    created_at           TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS Site_Access_ID (
    application_id          SERIAL PRIMARY KEY,
    full_name               VARCHAR(100) NOT NULL,
    job_title               VARCHAR(100),
    employee_number         VARCHAR(50),
    organization            VARCHAR(100) NOT NULL,
    subcontractor           VARCHAR(100),
    gov_id_type             VARCHAR(20),
    gov_id_number           VARCHAR(50),
    blood_type              VARCHAR(5),
    mobile_number           VARCHAR(20),
    whatsapp_number         VARCHAR(20),
    is_third_party_sponsor  BOOLEAN DEFAULT FALSE,
    workflow_state          VARCHAR(50) DEFAULT 'Pending Security Clearance',
    created_at              TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    submitted_by_user_id    INT REFERENCES App_Users(user_id)
);
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS email VARCHAR(150);
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS workflow_type VARCHAR(30) DEFAULT 'security_only';
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS form_variant VARCHAR(20) DEFAULT 'full';
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS date_joined DATE;
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS reports_to VARCHAR(100);
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS submitted_by_external_id INT;
ALTER TABLE Site_Access_ID ALTER COLUMN workflow_state SET DEFAULT 'Pending Security Clearance';
ALTER TABLE Site_Access_ID ALTER COLUMN job_title       DROP NOT NULL;
ALTER TABLE Site_Access_ID ALTER COLUMN employee_number DROP NOT NULL;
ALTER TABLE Site_Access_ID ALTER COLUMN gov_id_type     DROP NOT NULL;
ALTER TABLE Site_Access_ID ALTER COLUMN gov_id_number   DROP NOT NULL;
ALTER TABLE Site_Access_ID ALTER COLUMN blood_type      DROP NOT NULL;
ALTER TABLE Site_Access_ID ALTER COLUMN mobile_number   DROP NOT NULL;
ALTER TABLE Site_Access_ID ALTER COLUMN whatsapp_number DROP NOT NULL;
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS iqama_expiry_date DATE;
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS passport_expiry_date DATE;
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS insurance_expiry_date DATE;
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS gender VARCHAR(10);
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS nationality VARCHAR(100);
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS date_of_birth DATE;
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS passport_number VARCHAR(50);
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS language VARCHAR(50);
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS safety_induction_date DATE;
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS sponsor_name VARCHAR(150);

CREATE TABLE IF NOT EXISTS Non_Grata (
    ng_id SERIAL PRIMARY KEY,
    full_name VARCHAR(100) NOT NULL,
    gender VARCHAR(10),
    nationality VARCHAR(100),
    gov_id_type VARCHAR(20),
    gov_id_number VARCHAR(50),
    date_of_birth DATE,
    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    created_by INT REFERENCES App_Users(user_id)
);

CREATE TABLE IF NOT EXISTS Induction_Workbench (
    iw_id SERIAL PRIMARY KEY,
    full_name VARCHAR(150) NOT NULL,
    gender VARCHAR(10),
    nationality VARCHAR(100),
    date_of_birth DATE,
    gov_id_number VARCHAR(50),
    mobile_number VARCHAR(20),
    blood_type VARCHAR(10),
    language VARCHAR(50),
    company VARCHAR(200),
    legal_agreement VARCHAR(50),
    insurance_validity DATE,
    profession_iqama VARCHAR(200),
    profession_work VARCHAR(200),
    location VARCHAR(200),
    induction_date DATE,
    status VARCHAR(50),
    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    created_by INT REFERENCES App_Users(user_id)
);

CREATE TABLE IF NOT EXISTS System_Users (
    su_id SERIAL PRIMARY KEY,
    employee_name VARCHAR(150) NOT NULL,
    gov_id_number VARCHAR(50),
    email VARCHAR(200),
    project VARCHAR(200),
    company VARCHAR(200),
    job_title VARCHAR(200),
    location VARCHAR(200),
    mobile_number VARCHAR(20),
    nationality VARCHAR(100),
    role VARCHAR(100),
    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    created_by INT REFERENCES App_Users(user_id)
);

CREATE TABLE IF NOT EXISTS Workflow_History (
    history_id            SERIAL PRIMARY KEY,
    application_id        INT NOT NULL,
    action_taken          VARCHAR(50) NOT NULL,
    previous_state        VARCHAR(50),
    new_state             VARCHAR(50) NOT NULL,
    acted_by_user_id      INT REFERENCES App_Users(user_id),
    action_timestamp      TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    comments              TEXT
);

CREATE TABLE IF NOT EXISTS Roles (
    role_id      SERIAL PRIMARY KEY,
    role_name    VARCHAR(50) UNIQUE NOT NULL,
    description  VARCHAR(255)
);

CREATE TABLE IF NOT EXISTS Role_Permissions (
    role_id          INT NOT NULL REFERENCES Roles(role_id) ON DELETE CASCADE,
    permission_key   VARCHAR(50) NOT NULL,
    PRIMARY KEY (role_id, permission_key)
);

CREATE TABLE IF NOT EXISTS Attachments (
    attachment_id        SERIAL PRIMARY KEY,
    application_id       INT NOT NULL REFERENCES Site_Access_ID(application_id) ON DELETE CASCADE,
    attachment_type      VARCHAR(50) NOT NULL,
    file_path            VARCHAR(500) NOT NULL,
    original_filename    VARCHAR(255) NOT NULL,
    mime_type            VARCHAR(100),
    file_size_bytes      BIGINT,
    uploaded_at          TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    uploaded_by_user_id  INT REFERENCES App_Users(user_id)
);

CREATE TABLE IF NOT EXISTS Notifications (
    notification_id    SERIAL PRIMARY KEY,
    application_id     INT REFERENCES Site_Access_ID(application_id) ON DELETE CASCADE,
    channel            VARCHAR(20) NOT NULL,
    recipient          VARCHAR(150) NOT NULL,
    subject            VARCHAR(255),
    body               TEXT NOT NULL,
    delivery_status    VARCHAR(30) NOT NULL,
    error_message      TEXT,
    created_at         TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    sent_at            TIMESTAMP WITH TIME ZONE
);

CREATE TABLE IF NOT EXISTS Groups (
    group_id      SERIAL PRIMARY KEY,
    group_name    VARCHAR(80) UNIQUE NOT NULL,
    description   VARCHAR(255),
    created_at    TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS User_Group_Members (
    user_id    INT NOT NULL REFERENCES App_Users(user_id)  ON DELETE CASCADE,
    group_id   INT NOT NULL REFERENCES Groups(group_id)    ON DELETE CASCADE,
    added_at   TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    PRIMARY KEY (user_id, group_id)
);

CREATE TABLE IF NOT EXISTS Submitter_Organizations (
    organization_id   SERIAL PRIMARY KEY,
    organization_name VARCHAR(100) UNIQUE NOT NULL,
    submitter_group   VARCHAR(20) NOT NULL,
    description       VARCHAR(255),
    is_active         BOOLEAN DEFAULT TRUE,
    created_at        TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS External_Submitters (
    submitter_id        SERIAL PRIMARY KEY,
    full_name           VARCHAR(100) NOT NULL,
    email               VARCHAR(150) UNIQUE NOT NULL,
    organization_id     INT NOT NULL REFERENCES Submitter_Organizations(organization_id),
    hashed_password     VARCHAR(255) NOT NULL,
    must_change_password BOOLEAN DEFAULT TRUE,
    is_active           BOOLEAN DEFAULT TRUE,
    last_login_at       TIMESTAMP WITH TIME ZONE,
    created_at          TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    created_by_user_id  INT REFERENCES App_Users(user_id)
);

CREATE TABLE IF NOT EXISTS System_Settings (
    setting_key   VARCHAR(50) PRIMARY KEY,
    setting_value VARCHAR(255) NOT NULL,
    description   VARCHAR(255),
    updated_at    TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    updated_by_user_id INT REFERENCES App_Users(user_id)
);

-- v0.8 NEW
CREATE TABLE IF NOT EXISTS Sessions (
    session_token   VARCHAR(64) PRIMARY KEY,
    principal_type  VARCHAR(20) NOT NULL,           -- 'staff' or 'external'
    principal_id    INT          NOT NULL,
    issued_at       TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    expires_at      TIMESTAMP WITH TIME ZONE NOT NULL,
    last_seen_at    TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    user_agent      VARCHAR(255),
    ip_address      VARCHAR(45)
);
CREATE INDEX IF NOT EXISTS idx_sessions_expires_at ON Sessions(expires_at);

CREATE TABLE IF NOT EXISTS Failed_Login_Attempts (
    attempt_id      SERIAL PRIMARY KEY,
    principal_type  VARCHAR(20) NOT NULL,
    identifier      VARCHAR(150) NOT NULL,          -- username or email
    ip_address      VARCHAR(45),
    attempted_at    TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_failed_login_identifier_ip
    ON Failed_Login_Attempts(identifier, ip_address, attempted_at);

-- v0.9 NEW: Configurable lookup tables for dropdowns
CREATE TABLE IF NOT EXISTS Lookup_Tables (
    table_id      SERIAL PRIMARY KEY,
    name          VARCHAR(50) UNIQUE NOT NULL,
    description   VARCHAR(255),
    created_at    TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS Lookup_Values (
    value_id         SERIAL PRIMARY KEY,
    table_id         INT NOT NULL REFERENCES Lookup_Tables(table_id) ON DELETE CASCADE,
    value            VARCHAR(255) NOT NULL,
    label            VARCHAR(255),
    display_order    INT DEFAULT 0,
    is_active        BOOLEAN DEFAULT TRUE,
    created_at       TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(table_id, value)
);
"""

PERMISSION_CATALOG = [
    ("submit_application",     "Can submit a new ID application"),
    ("approve_security",       "Can approve/reject at Security stage"),
    ("approve_hr",             "Can approve/reject at HR stage (legacy / standard workflow)"),
    ("view_all_applications",  "Can see every application"),
    ("edit_application",       "Can edit application fields while still pending"),
    ("manage_users",           "Can administer staff users, roles, and groups"),
    ("manage_external_users",  "Can create/manage external submitter accounts"),
    ("manage_system_settings", "Can change system settings (workflow mode, lookup tables, etc.)"),
]


SEED_SQL = """
INSERT INTO Roles (role_id, role_name, description) VALUES
    (1, 'Submitter',          'Internal staff who can submit applications'),
    (2, 'HR Verification',    'HR approver (used in standard workflow)'),
    (3, 'Security Clearance', 'Approves applications at the Security stage'),
    (4, 'Reviewer',           'General reviewer (read-only)'),
    (5, 'Administrator',      'Full system administrator')
ON CONFLICT (role_id) DO NOTHING;

SELECT setval(pg_get_serial_sequence('roles','role_id'),
              GREATEST((SELECT MAX(role_id) FROM Roles), 1));

INSERT INTO Role_Permissions (role_id, permission_key) VALUES
    (1, 'submit_application'),
    (2, 'approve_hr'),
    (2, 'view_all_applications'),
    (3, 'approve_security'),
    (3, 'view_all_applications'),
    (3, 'edit_application'),
    (4, 'view_all_applications'),
    (5, 'submit_application'),
    (5, 'approve_security'),
    (5, 'approve_hr'),
    (5, 'view_all_applications'),
    (5, 'edit_application'),
    (5, 'manage_users'),
    (5, 'manage_external_users'),
    (5, 'manage_system_settings')
ON CONFLICT (role_id, permission_key) DO NOTHING;

INSERT INTO App_Users (full_name, username, hashed_password, role_id, force_password_change) VALUES
    ('Omar Abdulaziz Almaqhawi','omar.almaqhawi',        'temp_hash_123', 5, TRUE),
    ('Abdulkader Abdullatif',   'abdulkader.abdullatif', 'temp_hash_123', 3, TRUE),
    ('Hind Mana Alahmari',      'hind.alahmari',         'temp_hash_123', 4, TRUE),
    ('Badriah Rizq Allah',      'badriah.rizqallah',     'temp_hash_123', 1, TRUE)
ON CONFLICT (username) DO NOTHING;

INSERT INTO Submitter_Organizations (organization_name, submitter_group, description) VALUES
    ('SEVEN',                'group_1', 'Client'),
    ('ELLISDON',             'group_1', 'Project Management Consultant (PMC)'),
    ('SBG',                  'group_1', 'Main Contractor'),
    ('AECOM',                'group_1', 'Lead Design Consultant (LDC)'),
    ('ATKINS',               'group_1', 'Lead Design Consultant (LDC)'),
    ('Subcontractor (other)','group_2', 'Catch-all for any subcontractor not separately listed')
ON CONFLICT (organization_name) DO NOTHING;

INSERT INTO System_Settings (setting_key, setting_value, description) VALUES
    ('workflow_mode', 'security_only',
     'security_only = single Security approval; standard = HR then Security')
ON CONFLICT (setting_key) DO NOTHING;

-- v0.9 NEW: Seed lookup tables for configurable dropdowns
INSERT INTO Lookup_Tables (name, description) VALUES
    ('companies', 'Employer / Company Name dropdown values'),
    ('subcontractors', 'Subcontractor name dropdown values')
ON CONFLICT (name) DO NOTHING;

-- Seed companies from the original Submitter_Organizations (Group 1 + Group 2 subcontractor)
WITH co AS (SELECT table_id FROM Lookup_Tables WHERE name = 'companies')
INSERT INTO Lookup_Values (table_id, value, label, display_order, is_active)
SELECT co.table_id, org, org, ROW_NUMBER() OVER (), TRUE
FROM (VALUES ('SEVEN'), ('ELLISDON'), ('SBG'), ('AECOM'), ('ATKINS'),
             ('Subcontractor (other)')) AS v(org)
CROSS JOIN co
ON CONFLICT (table_id, value) DO NOTHING;

-- Seed empty subcontractors lookup (admin populates via UI or Excel import)
INSERT INTO Lookup_Values (table_id, value, label, display_order, is_active)
SELECT table_id, 'placeholder', 'placeholder', 0, FALSE
FROM Lookup_Tables WHERE name = 'subcontractors'
ON CONFLICT DO NOTHING;
"""

MIGRATION_SQL = """
UPDATE App_Users
   SET role_id = 5
 WHERE full_name = 'Omar Abdulaziz Almaqhawi'
   AND role_id <> 5;

WITH affected AS (
    SELECT application_id FROM Site_Access_ID
     WHERE workflow_state = 'Pending HR Verification'
)
INSERT INTO Workflow_History
    (application_id, action_taken, previous_state, new_state, acted_by_user_id, comments)
SELECT application_id, 'Workflow Migration', 'Pending HR Verification',
       'Pending Security Clearance', NULL,
       'Auto-migrated to security-only workflow'
  FROM affected;

UPDATE Site_Access_ID
   SET workflow_state = 'Pending Security Clearance'
 WHERE workflow_state = 'Pending HR Verification';
"""


def hash_password(plaintext: str) -> str:
    return bcrypt.hashpw(plaintext.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plaintext: str, hashed: str) -> bool:
    if not hashed:
        return False
    try:
        return bcrypt.checkpw(plaintext.encode("utf-8"), hashed.encode("utf-8"))
    except (ValueError, TypeError):
        return False


def generate_password() -> str:
    return secrets.token_urlsafe(9)


@app.post("/api/v1/admin/init-db")
async def init_db(request: Request):
    # Safety: if admin users already exist, require auth + manage_system_settings.
    # On a fresh DB (no App_Users table yet), unauth calls are allowed.
    try:
        conn = await get_conn()
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": f"Database connection failed: {str(e)}", "hint": "Check that DATABASE_URL is set in Railway Variables."}
        )
    try:
        existing_admin = await conn.fetchval("SELECT 1 FROM App_Users LIMIT 1")
    except Exception:
        existing_admin = None  # table doesn't exist — fresh DB
    if existing_admin:
        await require_permission(request, "manage_system_settings")
    try:
        await conn.execute(SCHEMA_SQL)
        await conn.execute(SEED_SQL)
        await conn.execute(MIGRATION_SQL)
        for key, _desc in PERMISSION_CATALOG:
            await conn.execute(
                """INSERT INTO Role_Permissions (role_id, permission_key)
                   VALUES (5, $1) ON CONFLICT DO NOTHING""",
                key,
            )
        placeholder_users = await conn.fetch(
            "SELECT user_id FROM App_Users WHERE hashed_password = 'temp_hash_123'"
        )
        if placeholder_users:
            real_hash = hash_password(DEFAULT_SEED_PASSWORD)
            for u in placeholder_users:
                await conn.execute(
                    """UPDATE App_Users
                          SET hashed_password = $1, force_password_change = TRUE
                        WHERE user_id = $2""",
                    real_hash, u["user_id"],
                )
        return {
            "status": "ok",
            "message": "Schema, seeds, and migrations applied.",
            "default_seed_password_hint": (
                f"Seeded staff users now log in with '{DEFAULT_SEED_PASSWORD}' "
                "and will be forced to change it."
            ),
        }
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": f"SQL execution failed: {str(e)}"}
        )
    finally:
        await conn.close()


# ============================================================
#  Models
# ============================================================
class LoginRequest(BaseModel):
    username: str
    password: str


class ExternalLoginRequest(BaseModel):
    email: str
    password: str


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class StatusUpdate(BaseModel):
    new_state: str
    comments: Optional[str] = None


class UserCreate(BaseModel):
    full_name: str
    username: str
    role_id: int
    is_active: bool = True


class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    role_id: Optional[int] = None
    is_active: Optional[bool] = None


class RoleCreate(BaseModel):
    role_name: str
    description: Optional[str] = None
    permissions: List[str] = []


class RolePermissionsUpdate(BaseModel):
    permissions: List[str]


class GroupCreate(BaseModel):
    group_name: str
    description: Optional[str] = None
    member_user_ids: List[int] = []


class GroupMembersUpdate(BaseModel):
    member_user_ids: List[int]


class ExternalSubmitterCreate(BaseModel):
    full_name: str
    email: str
    organization_id: int


class ApplicationFieldsUpdate(BaseModel):
    full_name:        Optional[str] = None
    job_title:        Optional[str] = None
    employee_number:  Optional[str] = None
    organization:     Optional[str] = None
    subcontractor:    Optional[str] = None
    gov_id_type:      Optional[str] = None
    gov_id_number:    Optional[str] = None
    blood_type:       Optional[str] = None
    mobile_number:    Optional[str] = None
    whatsapp_number:  Optional[str] = None
    email:            Optional[str] = None
    is_third_party_sponsor: Optional[bool] = None
    date_joined:      Optional[str] = None
    reports_to:       Optional[str] = None
    iqama_expiry_date:   Optional[str] = None
    passport_expiry_date: Optional[str] = None
    insurance_expiry_date: Optional[str] = None
    gender: Optional[str] = None
    nationality: Optional[str] = None
    date_of_birth: Optional[str] = None
    passport_number: Optional[str] = None
    language: Optional[str] = None
    safety_induction_date: Optional[str] = None
    sponsor_name: Optional[str] = None


class SettingUpdate(BaseModel):
    setting_value: str


class LookupValueCreate(BaseModel):
    value: str
    label: Optional[str] = None
    display_order: Optional[int] = 0
    is_active: Optional[bool] = True


class LookupValueUpdate(BaseModel):
    value: Optional[str] = None
    label: Optional[str] = None
    display_order: Optional[int] = None
    is_active: Optional[bool] = None


# ============================================================
#  Session helpers
# ============================================================
async def _create_session(conn, principal_type: str, principal_id: int,
                          user_agent: Optional[str], ip: Optional[str]) -> Tuple[str, datetime]:
    token = secrets.token_urlsafe(48)
    expires_at = datetime.utcnow() + SESSION_DURATION
    await conn.execute(
        """INSERT INTO Sessions
              (session_token, principal_type, principal_id, expires_at, user_agent, ip_address)
           VALUES ($1, $2, $3, $4, $5, $6)""",
        token, principal_type, principal_id, expires_at,
        (user_agent or "")[:255], (ip or "")[:45],
    )
    return token, expires_at


async def _cleanup_expired_sessions(conn):
    await conn.execute("DELETE FROM Sessions WHERE expires_at < NOW()")


async def _revoke_session(conn, token: str):
    await conn.execute("DELETE FROM Sessions WHERE session_token = $1", token)


async def _refresh_session(conn, token: str):
    new_expiry = datetime.utcnow() + SESSION_DURATION
    await conn.execute(
        """UPDATE Sessions
              SET last_seen_at = NOW(), expires_at = $1
            WHERE session_token = $2""",
        new_expiry, token,
    )
    return new_expiry


async def _get_session(conn, token: str):
    if not token:
        return None
    return await conn.fetchrow(
        """SELECT session_token, principal_type, principal_id, expires_at
             FROM Sessions WHERE session_token = $1 AND expires_at > NOW()""",
        token,
    )


def _set_session_cookie(response: Response, token: str, expires_at: datetime):
    response.set_cookie(
        key=SESSION_COOKIE_NAME,
        value=token,
        max_age=int(SESSION_DURATION.total_seconds()),
        httponly=True,
        secure=True,
        samesite="none",   # cross-site frontend->backend on Railway requires None
        path="/",
    )


def _clear_session_cookie(response: Response):
    response.delete_cookie(
        key=SESSION_COOKIE_NAME,
        path="/",
        secure=True,
        samesite="none",
    )


# ============================================================
#  Throttling helpers
# ============================================================
async def _record_failed_login(conn, principal_type: str, identifier: str, ip: Optional[str]):
    await conn.execute(
        """INSERT INTO Failed_Login_Attempts (principal_type, identifier, ip_address)
           VALUES ($1, $2, $3)""",
        principal_type, identifier[:150], (ip or "")[:45],
    )


async def _is_throttled(conn, identifier: str, ip: Optional[str]) -> bool:
    cutoff = datetime.utcnow() - LOGIN_THROTTLE_WINDOW
    row = await conn.fetchrow(
        """SELECT COUNT(*) AS n
             FROM Failed_Login_Attempts
            WHERE identifier = $1
              AND attempted_at >= $2""",
        identifier[:150], cutoff,
    )
    return (row["n"] or 0) >= LOGIN_THROTTLE_MAX_ATTEMPTS


async def _clear_throttle(conn, identifier: str):
    await conn.execute(
        "DELETE FROM Failed_Login_Attempts WHERE identifier = $1",
        identifier[:150],
    )


# ============================================================
#  Auth: identity resolution dependency
# ============================================================
class Principal:
    """Resolved request identity. Either a staff user or an external submitter."""

    def __init__(self, principal_type: str, principal_id: int, full_name: str,
                 email: Optional[str], extras: dict, must_change_password: bool):
        self.principal_type = principal_type   # 'staff' or 'external'
        self.principal_id = principal_id
        self.full_name = full_name
        self.email = email
        self.extras = extras
        self.must_change_password = must_change_password

    @property
    def is_staff(self) -> bool:
        return self.principal_type == "staff"

    @property
    def is_external(self) -> bool:
        return self.principal_type == "external"


async def _resolve_principal(conn, principal_type: str, principal_id: int) -> Optional[Principal]:
    if principal_type == "staff":
        row = await conn.fetchrow(
            """SELECT u.user_id, u.full_name, u.username, u.role_id, u.is_active,
                      u.force_password_change, r.role_name
                 FROM App_Users u
            LEFT JOIN Roles r ON r.role_id = u.role_id
                WHERE u.user_id = $1""",
            principal_id,
        )
        if not row or not row["is_active"]:
            return None
        return Principal(
            "staff", row["user_id"], row["full_name"], None,
            {"username": row["username"], "role_id": row["role_id"], "role_name": row["role_name"]},
            row["force_password_change"],
        )
    elif principal_type == "external":
        row = await conn.fetchrow(
            """SELECT es.submitter_id, es.full_name, es.email, es.is_active,
                      es.must_change_password, es.organization_id,
                      so.organization_name, so.submitter_group
                 FROM External_Submitters es
            LEFT JOIN Submitter_Organizations so ON so.organization_id = es.organization_id
                WHERE es.submitter_id = $1""",
            principal_id,
        )
        if not row or not row["is_active"]:
            return None
        return Principal(
            "external", row["submitter_id"], row["full_name"], row["email"],
            {
                "organization_id": row["organization_id"],
                "organization_name": row["organization_name"],
                "submitter_group": row["submitter_group"],
            },
            row["must_change_password"],
        )
    return None


async def get_principal_or_none(request: Request) -> Optional[Principal]:
    """Returns the current Principal or None if not logged in."""
    token = request.cookies.get(SESSION_COOKIE_NAME)
    if not token:
        return None
    conn = await get_conn()
    try:
        sess = await _get_session(conn, token)
        if not sess:
            return None
        principal = await _resolve_principal(conn, sess["principal_type"], sess["principal_id"])
        if principal:
            await _refresh_session(conn, token)
        return principal
    finally:
        await conn.close()


async def require_login(request: Request) -> Principal:
    p = await get_principal_or_none(request)
    if not p:
        raise HTTPException(401, "Not logged in")
    return p


async def require_staff(request: Request) -> Principal:
    p = await require_login(request)
    if not p.is_staff:
        raise HTTPException(403, "Staff account required")
    if p.must_change_password:
        raise HTTPException(409, "Password change required before continuing")
    return p


async def require_external(request: Request) -> Principal:
    p = await require_login(request)
    if not p.is_external:
        raise HTTPException(403, "External submitter account required")
    if p.must_change_password:
        raise HTTPException(409, "Password change required before continuing")
    return p


async def require_permission(request: Request, permission: str) -> Principal:
    p = await require_staff(request)
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT rp.permission_key
                 FROM Role_Permissions rp
                 JOIN App_Users u ON u.role_id = rp.role_id
                WHERE u.user_id = $1""",
            p.principal_id,
        )
    finally:
        await conn.close()
    perms = {r["permission_key"] for r in rows}
    if permission not in perms:
        raise HTTPException(403, f"User lacks permission '{permission}'")
    return p


async def fetch_user_permissions(conn, user_id):
    rows = await conn.fetch(
        """SELECT rp.permission_key
             FROM Role_Permissions rp
             JOIN App_Users u ON u.role_id = rp.role_id
            WHERE u.user_id = $1""",
        user_id,
    )
    return [r["permission_key"] for r in rows]


async def get_workflow_mode(conn) -> str:
    row = await conn.fetchrow(
        "SELECT setting_value FROM System_Settings WHERE setting_key = 'workflow_mode'"
    )
    return row["setting_value"] if row else "security_only"


async def required_permission_for_state_change(conn, current_state: str, new_state: str) -> Optional[str]:
    mode = await get_workflow_mode(conn)
    if mode == "standard":
        if current_state == "Pending HR Verification" and new_state in ("Pending Security Clearance", "Rejected"):
            return "approve_hr"
        if current_state == "Pending Security Clearance" and new_state in ("Approved", "Rejected"):
            return "approve_security"
    else:
        if current_state == "Pending Security Clearance" and new_state in ("Approved", "Rejected"):
            return "approve_security"
        if current_state == "Pending HR Verification" and new_state in ("Pending Security Clearance", "Rejected"):
            return "approve_hr"
    return None


def parse_date_or_none(s: Optional[str]) -> Optional[date]:
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, f"Invalid date format '{s}' (expected YYYY-MM-DD)")


# ============================================================
#  Notifications (unchanged from v0.7)
# ============================================================
def _build_rejection_email_body(applicant_name, app_id, reason):
    return (
        f"Dear {applicant_name},\n\n"
        f"Your Site Access ID application (#{app_id}) has been reviewed and "
        f"unfortunately could not be approved at this time.\n\n"
        f"Reason provided by the reviewer:\n{reason}\n\n"
        f"If you believe this was made in error or you would like to re-apply with "
        f"corrected details, please contact your site administrator.\n\n"
        f"Regards,\nTa'awon P3 Security Management System"
    )


def _build_rejection_whatsapp_body(applicant_name, app_id, reason):
    return (
        f"Hello {applicant_name}, your Site Access ID application "
        f"(#{app_id}) was not approved.\nReason: {reason}\n"
        f"Please contact your site administrator for next steps."
    )


async def _log_notification(conn, app_id, channel, recipient, subject, body,
                            status, error_message=None, sent_at=None):
    await conn.execute(
        """INSERT INTO Notifications
              (application_id, channel, recipient, subject, body,
               delivery_status, error_message, sent_at)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8)""",
        app_id, channel, recipient, subject, body,
        status, error_message, sent_at,
    )


def _send_email_smtp(to_addr, subject, body):
    if not SMTP_ENABLED:
        return False, "SMTP not configured (SMTP_USERNAME/SMTP_PASSWORD missing in .env)"
    try:
        msg = MIMEMultipart()
        msg["From"] = SMTP_FROM
        msg["To"] = to_addr
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain", "utf-8"))
        context = ssl.create_default_context()
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as server:
            server.starttls(context=context)
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(SMTP_FROM, [to_addr], msg.as_string())
        return True, None
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


async def _handle_rejection_notifications(conn, app_record, reason):
    app_id = app_record["application_id"]
    name = app_record["full_name"]
    email = app_record["email"]
    whatsapp = app_record["whatsapp_number"] or app_record["mobile_number"]

    if not email:
        await _log_notification(conn, app_id, "email", "(unknown)", None,
                                "(skipped - no email on file)", "skipped_no_recipient")
    else:
        subject = f"Site Access ID Application #{app_id} - Update"
        body = _build_rejection_email_body(name, app_id, reason)
        success, err = _send_email_smtp(email, subject, body)
        await _log_notification(conn, app_id, "email", email, subject, body,
                                "sent" if success else "failed",
                                err, datetime.utcnow() if success else None)

    if not whatsapp:
        await _log_notification(conn, app_id, "whatsapp", "(unknown)", None,
                                "(skipped - no WhatsApp number)", "skipped_no_recipient")
    else:
        body = _build_rejection_whatsapp_body(name, app_id, reason)
        await _log_notification(conn, app_id, "whatsapp", whatsapp, None, body,
                                "pending_whatsapp_api")


# ============================================================
#  Utilities
# ============================================================
def _client_ip(request: Request) -> Optional[str]:
    fwd = request.headers.get("x-forwarded-for")
    if fwd:
        return fwd.split(",")[0].strip()
    return request.client.host if request.client else None


def _user_agent(request: Request) -> Optional[str]:
    return request.headers.get("user-agent")


# ============================================================
#  Auth endpoints
# ============================================================
@app.post("/api/v1/auth/login")
async def staff_login(payload: LoginRequest, request: Request, response: Response):
    username = payload.username.strip()
    if not username or not payload.password:
        raise HTTPException(400, "Username and password are required")
    conn = await get_conn()
    try:
        if await _is_throttled(conn, username, _client_ip(request)):
            raise HTTPException(
                429,
                f"Too many failed attempts. Please wait {LOGIN_THROTTLE_WINDOW.seconds // 60} minutes before trying again.",
            )
        user = await conn.fetchrow(
            """SELECT user_id, full_name, username, hashed_password, role_id,
                      is_active, force_password_change
                 FROM App_Users WHERE username = $1""",
            username,
        )
        if (not user) or (not user["is_active"]) or (not verify_password(payload.password, user["hashed_password"])):
            await _record_failed_login(conn, "staff", username, _client_ip(request))
            raise HTTPException(401, "Invalid username or password")
        await _clear_throttle(conn, username)
        await _cleanup_expired_sessions(conn)
        token, expires_at = await _create_session(
            conn, "staff", user["user_id"], _user_agent(request), _client_ip(request)
        )
        _set_session_cookie(response, token, expires_at)
        return {
            "status": "ok",
            "principal_type": "staff",
            "user_id": user["user_id"],
            "full_name": user["full_name"],
            "must_change_password": user["force_password_change"],
        }
    finally:
        await conn.close()


@app.post("/api/v1/auth/external-login")
async def external_login(payload: ExternalLoginRequest, request: Request, response: Response):
    email = payload.email.strip().lower()
    if not email or not payload.password:
        raise HTTPException(400, "Email and password are required")
    conn = await get_conn()
    try:
        if await _is_throttled(conn, email, _client_ip(request)):
            raise HTTPException(
                429,
                f"Too many failed attempts. Please wait {LOGIN_THROTTLE_WINDOW.seconds // 60} minutes before trying again.",
            )
        sub = await conn.fetchrow(
            """SELECT submitter_id, full_name, email, hashed_password, is_active,
                      must_change_password
                 FROM External_Submitters WHERE LOWER(email) = $1""",
            email,
        )
        if (not sub) or (not sub["is_active"]) or (not verify_password(payload.password, sub["hashed_password"])):
            await _record_failed_login(conn, "external", email, _client_ip(request))
            raise HTTPException(401, "Invalid email or password")
        await _clear_throttle(conn, email)
        await _cleanup_expired_sessions(conn)
        token, expires_at = await _create_session(
            conn, "external", sub["submitter_id"], _user_agent(request), _client_ip(request)
        )
        _set_session_cookie(response, token, expires_at)
        await conn.execute(
            "UPDATE External_Submitters SET last_login_at = NOW() WHERE submitter_id = $1",
            sub["submitter_id"],
        )
        return {
            "status": "ok",
            "principal_type": "external",
            "submitter_id": sub["submitter_id"],
            "full_name": sub["full_name"],
            "must_change_password": sub["must_change_password"],
        }
    finally:
        await conn.close()


@app.post("/api/v1/auth/logout")
async def logout(request: Request, response: Response):
    token = request.cookies.get(SESSION_COOKIE_NAME)
    if token:
        conn = await get_conn()
        try:
            await _revoke_session(conn, token)
        finally:
            await conn.close()
    _clear_session_cookie(response)
    return {"status": "ok"}


@app.get("/api/v1/auth/me")
async def whoami(request: Request):
    p = await get_principal_or_none(request)
    if not p:
        return {"logged_in": False}
    conn = await get_conn()
    try:
        permissions = []
        if p.is_staff:
            permissions = await fetch_user_permissions(conn, p.principal_id)
    finally:
        await conn.close()
    return {
        "logged_in": True,
        "principal_type": p.principal_type,
        "principal_id": p.principal_id,
        "full_name": p.full_name,
        "email": p.email,
        "must_change_password": p.must_change_password,
        "permissions": permissions,
        **p.extras,
    }


@app.post("/api/v1/auth/change-password")
async def change_password(payload: ChangePasswordRequest, request: Request):
    p = await require_login(request)
    if len(payload.new_password) < 8:
        raise HTTPException(400, "New password must be at least 8 characters")
    if payload.new_password == payload.current_password:
        raise HTTPException(400, "New password must be different from current password")
    conn = await get_conn()
    try:
        if p.is_staff:
            row = await conn.fetchrow(
                "SELECT hashed_password FROM App_Users WHERE user_id = $1", p.principal_id
            )
            if not row or not verify_password(payload.current_password, row["hashed_password"]):
                raise HTTPException(401, "Current password is incorrect")
            await conn.execute(
                """UPDATE App_Users
                      SET hashed_password = $1, force_password_change = FALSE
                    WHERE user_id = $2""",
                hash_password(payload.new_password), p.principal_id,
            )
        else:
            row = await conn.fetchrow(
                "SELECT hashed_password FROM External_Submitters WHERE submitter_id = $1",
                p.principal_id,
            )
            if not row or not verify_password(payload.current_password, row["hashed_password"]):
                raise HTTPException(401, "Current password is incorrect")
            await conn.execute(
                """UPDATE External_Submitters
                      SET hashed_password = $1, must_change_password = FALSE
                    WHERE submitter_id = $2""",
                hash_password(payload.new_password), p.principal_id,
            )
        return {"status": "ok"}
    finally:
        await conn.close()


# ============================================================
#  Public-ish endpoints (still require login)
# ============================================================
@app.get("/")
async def root():
    return {
        "service": "Ta'awon P3 SMS API",
        "version": app.version,
        "smtp_enabled": SMTP_ENABLED,
        "cors_origins": CORS_ORIGINS,
        "auth": "session-cookie",
    }


@app.get("/api/v1/health")
async def health_check():
    """Test database connectivity. Returns ok/error with details."""
    try:
        conn = await get_conn()
        try:
            ver = await conn.fetchval("SELECT version()")
            tables = await conn.fetch(
                "SELECT tablename FROM pg_tables WHERE schemaname = 'public' ORDER BY tablename"
            )
            return {
                "status": "ok",
                "database": ver.split(",")[0].strip(),
                "tables": [r["tablename"] for r in tables],
                "table_count": len(tables),
            }
        finally:
            await conn.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/api/v1/permissions")
async def list_permission_catalog(request: Request):
    await require_login(request)
    return {"permissions": [{"key": k, "description": d} for k, d in PERMISSION_CATALOG]}


@app.get("/api/v1/users")
async def list_users(request: Request):
    """Lists active staff users. Used by the dashboard for displaying acted_by names etc."""
    await require_staff(request)
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT u.user_id, u.full_name, u.username, u.role_id,
                      r.role_name, u.is_active
                 FROM App_Users u
            LEFT JOIN Roles r ON r.role_id = u.role_id
                WHERE u.is_active = TRUE
             ORDER BY u.user_id"""
        )
        return {"users": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.get("/api/v1/organizations")
async def list_organizations(request: Request):
    """For staff: all active organizations (used in admin/external-submitter creation).
       For external submitters: just their own organization."""
    p = await require_login(request)
    if p.must_change_password:
        raise HTTPException(409, "Password change required before continuing")
    conn = await get_conn()
    try:
        if p.is_external:
            rows = await conn.fetch(
                """SELECT organization_id, organization_name, submitter_group, description, is_active
                     FROM Submitter_Organizations
                    WHERE organization_id = $1""",
                p.extras["organization_id"],
            )
        else:
            rows = await conn.fetch(
                """SELECT organization_id, organization_name, submitter_group, description, is_active
                     FROM Submitter_Organizations
                    WHERE is_active = TRUE
                 ORDER BY submitter_group, organization_name"""
            )
        return {"organizations": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.get("/api/v1/settings")
async def public_settings(request: Request):
    await require_login(request)
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            "SELECT setting_key, setting_value FROM System_Settings WHERE setting_key IN ('workflow_mode')"
        )
        return {"settings": {r["setting_key"]: r["setting_value"] for r in rows}}
    finally:
        await conn.close()


# ============================================================
#  Admin: Users
# ============================================================
@app.get("/api/v1/admin/users")
async def admin_list_users(request: Request):
    await require_permission(request, "manage_users")
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT u.user_id, u.full_name, u.username, u.role_id, r.role_name,
                      u.is_active, u.force_password_change, u.created_at
                 FROM App_Users u
            LEFT JOIN Roles r ON r.role_id = u.role_id
             ORDER BY u.user_id"""
        )
        return {"users": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.post("/api/v1/admin/users")
async def admin_create_user(payload: UserCreate, request: Request):
    await require_permission(request, "manage_users")
    if not payload.full_name.strip() or not payload.username.strip():
        raise HTTPException(400, "Full name and username are required")
    plain_password = generate_password()
    hashed = hash_password(plain_password)
    conn = await get_conn()
    try:
        try:
            row = await conn.fetchrow(
                """INSERT INTO App_Users (full_name, username, hashed_password, role_id, is_active, force_password_change)
                   VALUES ($1, $2, $3, $4, $5, TRUE)
                RETURNING user_id, full_name, username, role_id, is_active""",
                payload.full_name.strip(), payload.username.strip(),
                hashed, payload.role_id, payload.is_active,
            )
        except asyncpg.UniqueViolationError:
            raise HTTPException(409, f"Username '{payload.username}' already exists")
        return {
            "status": "ok",
            "user": dict(row),
            "temporary_password": plain_password,
            "note": "Share this password with the user. They will be required to change it on first login. This password will NOT be shown again.",
        }
    finally:
        await conn.close()


@app.put("/api/v1/admin/users/{user_id}")
async def admin_update_user(user_id: int, payload: UserUpdate, request: Request):
    await require_permission(request, "manage_users")
    sets, vals = [], []
    if payload.full_name is not None:
        sets.append(f"full_name = ${len(vals) + 1}"); vals.append(payload.full_name.strip())
    if payload.role_id is not None:
        sets.append(f"role_id = ${len(vals) + 1}"); vals.append(payload.role_id)
    if payload.is_active is not None:
        sets.append(f"is_active = ${len(vals) + 1}"); vals.append(payload.is_active)
    if not sets:
        raise HTTPException(400, "No fields to update")
    vals.append(user_id)
    conn = await get_conn()
    try:
        q = f"UPDATE App_Users SET {', '.join(sets)} WHERE user_id = ${len(vals)} RETURNING user_id"
        row = await conn.fetchrow(q, *vals)
        if not row:
            raise HTTPException(404, "User not found")
        return {"status": "ok", "user_id": row["user_id"]}
    finally:
        await conn.close()


@app.put("/api/v1/admin/users/{user_id}/regenerate-password")
async def admin_regenerate_staff_password(user_id: int, request: Request):
    p = await require_permission(request, "manage_users")
    plain_password = generate_password()
    hashed = hash_password(plain_password)
    conn = await get_conn()
    try:
        row = await conn.fetchrow(
            """UPDATE App_Users
                  SET hashed_password = $1, force_password_change = TRUE
                WHERE user_id = $2
            RETURNING user_id, full_name""",
            hashed, user_id,
        )
        if not row:
            raise HTTPException(404, "User not found")
        # Also revoke any existing sessions for this user
        await conn.execute(
            "DELETE FROM Sessions WHERE principal_type = 'staff' AND principal_id = $1",
            user_id,
        )
        return {
            "status": "ok",
            "user": dict(row),
            "temporary_password": plain_password,
            "note": "Share this password with the user. They will be required to change it on next login.",
        }
    finally:
        await conn.close()


@app.delete("/api/v1/admin/users/{user_id}")
async def admin_deactivate_user(user_id: int, request: Request):
    p = await require_permission(request, "manage_users")
    if p.principal_id == user_id:
        raise HTTPException(400, "You cannot deactivate yourself")
    conn = await get_conn()
    try:
        row = await conn.fetchrow(
            "UPDATE App_Users SET is_active = FALSE WHERE user_id = $1 RETURNING user_id",
            user_id,
        )
        if not row:
            raise HTTPException(404, "User not found")
        await conn.execute(
            "DELETE FROM Sessions WHERE principal_type = 'staff' AND principal_id = $1",
            user_id,
        )
        return {"status": "ok"}
    finally:
        await conn.close()


# ============================================================
#  Admin: Roles & Permissions
# ============================================================
@app.get("/api/v1/admin/roles")
async def admin_list_roles(request: Request):
    await require_permission(request, "manage_users")
    conn = await get_conn()
    try:
        roles = await conn.fetch("SELECT role_id, role_name, description FROM Roles ORDER BY role_id")
        out = []
        for r in roles:
            perms = await conn.fetch(
                "SELECT permission_key FROM Role_Permissions WHERE role_id = $1 ORDER BY permission_key",
                r["role_id"],
            )
            out.append({**dict(r), "permissions": [p["permission_key"] for p in perms]})
        return {"roles": out}
    finally:
        await conn.close()


@app.post("/api/v1/admin/roles")
async def admin_create_role(payload: RoleCreate, request: Request):
    await require_permission(request, "manage_users")
    valid_keys = {k for k, _ in PERMISSION_CATALOG}
    bad = [p for p in payload.permissions if p not in valid_keys]
    if bad:
        raise HTTPException(400, f"Unknown permission keys: {bad}")
    conn = await get_conn()
    try:
        async with conn.transaction():
            try:
                row = await conn.fetchrow(
                    """INSERT INTO Roles (role_name, description) VALUES ($1, $2)
                    RETURNING role_id, role_name, description""",
                    payload.role_name.strip(), payload.description,
                )
            except asyncpg.UniqueViolationError:
                raise HTTPException(409, f"Role '{payload.role_name}' already exists")
            for key in payload.permissions:
                await conn.execute(
                    "INSERT INTO Role_Permissions (role_id, permission_key) VALUES ($1, $2)",
                    row["role_id"], key,
                )
        return {"status": "ok", "role": dict(row), "permissions": payload.permissions}
    finally:
        await conn.close()


@app.put("/api/v1/admin/roles/{role_id}/permissions")
async def admin_update_role_permissions(role_id: int, payload: RolePermissionsUpdate, request: Request):
    await require_permission(request, "manage_users")
    valid_keys = {k for k, _ in PERMISSION_CATALOG}
    bad = [p for p in payload.permissions if p not in valid_keys]
    if bad:
        raise HTTPException(400, f"Unknown permission keys: {bad}")
    conn = await get_conn()
    try:
        role = await conn.fetchrow("SELECT role_id FROM Roles WHERE role_id = $1", role_id)
        if not role:
            raise HTTPException(404, "Role not found")
        async with conn.transaction():
            await conn.execute("DELETE FROM Role_Permissions WHERE role_id = $1", role_id)
            for key in payload.permissions:
                await conn.execute(
                    "INSERT INTO Role_Permissions (role_id, permission_key) VALUES ($1, $2)",
                    role_id, key,
                )
        return {"status": "ok"}
    finally:
        await conn.close()


# ============================================================
#  Admin: Groups
# ============================================================
@app.get("/api/v1/admin/groups")
async def admin_list_groups(request: Request):
    await require_permission(request, "manage_users")
    conn = await get_conn()
    try:
        groups = await conn.fetch("SELECT group_id, group_name, description, created_at FROM Groups ORDER BY group_id")
        out = []
        for g in groups:
            members = await conn.fetch(
                """SELECT u.user_id, u.full_name, u.username
                     FROM User_Group_Members m
                     JOIN App_Users u ON u.user_id = m.user_id
                    WHERE m.group_id = $1
                 ORDER BY u.full_name""",
                g["group_id"],
            )
            out.append({**dict(g), "members": [dict(m) for m in members]})
        return {"groups": out}
    finally:
        await conn.close()


@app.post("/api/v1/admin/groups")
async def admin_create_group(payload: GroupCreate, request: Request):
    await require_permission(request, "manage_users")
    conn = await get_conn()
    try:
        async with conn.transaction():
            try:
                row = await conn.fetchrow(
                    """INSERT INTO Groups (group_name, description) VALUES ($1, $2)
                    RETURNING group_id, group_name, description""",
                    payload.group_name.strip(), payload.description,
                )
            except asyncpg.UniqueViolationError:
                raise HTTPException(409, f"Group '{payload.group_name}' already exists")
            for uid in payload.member_user_ids:
                await conn.execute(
                    "INSERT INTO User_Group_Members (user_id, group_id) VALUES ($1, $2)",
                    uid, row["group_id"],
                )
        return {"status": "ok", "group": dict(row)}
    finally:
        await conn.close()


@app.put("/api/v1/admin/groups/{group_id}/members")
async def admin_update_group_members(group_id: int, payload: GroupMembersUpdate, request: Request):
    await require_permission(request, "manage_users")
    conn = await get_conn()
    try:
        g = await conn.fetchrow("SELECT group_id FROM Groups WHERE group_id = $1", group_id)
        if not g:
            raise HTTPException(404, "Group not found")
        async with conn.transaction():
            await conn.execute("DELETE FROM User_Group_Members WHERE group_id = $1", group_id)
            for uid in payload.member_user_ids:
                await conn.execute(
                    "INSERT INTO User_Group_Members (user_id, group_id) VALUES ($1, $2)",
                    uid, group_id,
                )
        return {"status": "ok"}
    finally:
        await conn.close()


@app.delete("/api/v1/admin/groups/{group_id}")
async def admin_delete_group(group_id: int, request: Request):
    await require_permission(request, "manage_users")
    conn = await get_conn()
    try:
        row = await conn.fetchrow("DELETE FROM Groups WHERE group_id = $1 RETURNING group_id", group_id)
        if not row:
            raise HTTPException(404, "Group not found")
        return {"status": "ok"}
    finally:
        await conn.close()


# ============================================================
#  Admin: External Submitters
# ============================================================
@app.get("/api/v1/admin/external-submitters")
async def admin_list_external_submitters(request: Request):
    await require_permission(request, "manage_external_users")
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT es.submitter_id, es.full_name, es.email, es.is_active,
                      es.must_change_password, es.last_login_at, es.created_at,
                      so.organization_id, so.organization_name, so.submitter_group
                 FROM External_Submitters es
            LEFT JOIN Submitter_Organizations so ON so.organization_id = es.organization_id
             ORDER BY es.submitter_id DESC"""
        )
        return {"external_submitters": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.post("/api/v1/admin/external-submitters")
async def admin_create_external_submitter(payload: ExternalSubmitterCreate, request: Request):
    p = await require_permission(request, "manage_external_users")
    conn = await get_conn()
    try:
        org = await conn.fetchrow(
            "SELECT organization_id, organization_name FROM Submitter_Organizations WHERE organization_id = $1",
            payload.organization_id,
        )
        if not org:
            raise HTTPException(400, "Unknown organization_id")
        plain_password = generate_password()
        hashed = hash_password(plain_password)
        try:
            row = await conn.fetchrow(
                """INSERT INTO External_Submitters
                      (full_name, email, organization_id, hashed_password, created_by_user_id)
                   VALUES ($1, $2, $3, $4, $5)
                RETURNING submitter_id, full_name, email, organization_id""",
                payload.full_name.strip(), payload.email.strip(),
                payload.organization_id, hashed, p.principal_id,
            )
        except asyncpg.UniqueViolationError:
            raise HTTPException(409, f"An external submitter with email '{payload.email}' already exists")
        return {
            "status": "ok",
            "external_submitter": dict(row),
            "temporary_password": plain_password,
            "note": "Share this password with the user via WhatsApp or another secure channel. They will be required to change it on first login.",
        }
    finally:
        await conn.close()


@app.put("/api/v1/admin/external-submitters/{submitter_id}/regenerate-password")
async def admin_regenerate_external_password(submitter_id: int, request: Request):
    await require_permission(request, "manage_external_users")
    conn = await get_conn()
    try:
        sub = await conn.fetchrow(
            "SELECT submitter_id, full_name, email FROM External_Submitters WHERE submitter_id = $1",
            submitter_id,
        )
        if not sub:
            raise HTTPException(404, "External submitter not found")
        plain_password = generate_password()
        hashed = hash_password(plain_password)
        await conn.execute(
            """UPDATE External_Submitters
                  SET hashed_password = $1, must_change_password = TRUE
                WHERE submitter_id = $2""",
            hashed, submitter_id,
        )
        await conn.execute(
            "DELETE FROM Sessions WHERE principal_type = 'external' AND principal_id = $1",
            submitter_id,
        )
        return {
            "status": "ok",
            "submitter": dict(sub),
            "temporary_password": plain_password,
            "note": "Share with the user. Forces password change on next login.",
        }
    finally:
        await conn.close()


@app.delete("/api/v1/admin/external-submitters/{submitter_id}")
async def admin_deactivate_external_submitter(submitter_id: int, request: Request):
    await require_permission(request, "manage_external_users")
    conn = await get_conn()
    try:
        row = await conn.fetchrow(
            "UPDATE External_Submitters SET is_active = FALSE WHERE submitter_id = $1 RETURNING submitter_id",
            submitter_id,
        )
        if not row:
            raise HTTPException(404, "External submitter not found")
        await conn.execute(
            "DELETE FROM Sessions WHERE principal_type = 'external' AND principal_id = $1",
            submitter_id,
        )
        return {"status": "ok"}
    finally:
        await conn.close()


# ============================================================
#  Admin: System Settings
# ============================================================
@app.get("/api/v1/admin/settings")
async def admin_list_settings(request: Request):
    await require_permission(request, "manage_system_settings")
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            "SELECT setting_key, setting_value, description, updated_at FROM System_Settings ORDER BY setting_key"
        )
        return {"settings": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.put("/api/v1/admin/settings/{setting_key}")
async def admin_update_setting(setting_key: str, payload: SettingUpdate, request: Request):
    p = await require_permission(request, "manage_system_settings")
    if setting_key == "workflow_mode" and payload.setting_value not in ("security_only", "standard"):
        raise HTTPException(400, "workflow_mode must be 'security_only' or 'standard'")
    conn = await get_conn()
    try:
        row = await conn.fetchrow(
            """UPDATE System_Settings
                  SET setting_value = $1, updated_at = NOW(), updated_by_user_id = $2
                WHERE setting_key = $3
            RETURNING setting_key, setting_value""",
            payload.setting_value, p.principal_id, setting_key,
        )
        if not row:
            raise HTTPException(404, f"Unknown setting '{setting_key}'")
        return {"status": "ok", **dict(row)}
    finally:
        await conn.close()


# ============================================================
#  Lookup Tables (v0.9 — configurable dropdowns)
# ============================================================
@app.get("/api/v1/lookups")
async def list_all_lookups(request: Request):
    """Public endpoint — returns all lookup tables with their active values.
    Used by the frontend to populate form dropdowns."""
    conn = await get_conn()
    try:
        tables = await conn.fetch(
            "SELECT table_id, name, description FROM Lookup_Tables ORDER BY name"
        )
        result = []
        for t in tables:
            values = await conn.fetch(
                """SELECT value_id, value, label, display_order, is_active
                     FROM Lookup_Values
                    WHERE table_id = $1 AND is_active = TRUE
                 ORDER BY display_order ASC""",
                t["table_id"],
            )
            result.append({
                "name": t["name"],
                "description": t["description"],
                "values": [dict(v) for v in values],
            })
        return {"lookups": result}
    finally:
        await conn.close()


@app.get("/api/v1/lookups/{table_name}")
async def get_lookup(table_name: str, request: Request):
    """Get a single lookup table's values."""
    conn = await get_conn()
    try:
        t = await conn.fetchrow(
            "SELECT table_id, name, description FROM Lookup_Tables WHERE name = $1", table_name
        )
        if not t:
            raise HTTPException(404, f"Lookup table '{table_name}' not found")
        values = await conn.fetch(
            """SELECT value_id, value, label, display_order, is_active
                 FROM Lookup_Values
                WHERE table_id = $1 AND is_active = TRUE
             ORDER BY display_order ASC""",
            t["table_id"],
        )
        return {
            "name": t["name"],
            "description": t["description"],
            "values": [dict(v) for v in values],
        }
    finally:
        await conn.close()


@app.post("/api/v1/admin/lookups")
async def admin_create_lookup(payload: BaseModel, request: Request):
    """Create a new lookup table. Body: {"name": "...", "description": "..."}"""
    p = await require_permission(request, "manage_system_settings")
    data = await request.json()
    name = data.get("name", "").strip()
    description = data.get("description", "").strip()
    if not name or not name.replace("_", "").isalnum():
        raise HTTPException(400, "Name must be alphanumeric/underscores, non-empty")
    conn = await get_conn()
    try:
        row = await conn.fetchrow(
            """INSERT INTO Lookup_Tables (name, description)
               VALUES ($1, $2) ON CONFLICT (name) DO NOTHING
            RETURNING table_id, name, description""",
            name, description,
        )
        if not row:
            raise HTTPException(409, f"Lookup table '{name}' already exists")
        return {"status": "ok", "lookup": dict(row)}
    finally:
        await conn.close()


@app.post("/api/v1/admin/lookups/{table_name}/values")
async def admin_create_lookup_value(table_name: str, payload: LookupValueCreate, request: Request):
    p = await require_permission(request, "manage_system_settings")
    conn = await get_conn()
    try:
        t = await conn.fetchrow(
            "SELECT table_id FROM Lookup_Tables WHERE name = $1", table_name
        )
        if not t:
            raise HTTPException(404, f"Lookup table '{table_name}' not found")
        row = await conn.fetchrow(
            """INSERT INTO Lookup_Values (table_id, value, label, display_order, is_active)
               VALUES ($1, $2, $3, $4, $5)
            RETURNING value_id, value, label, display_order, is_active""",
            t["table_id"], payload.value, payload.label or payload.value,
            payload.display_order, payload.is_active,
        )
        return {"status": "ok", "value": dict(row)}
    finally:
        await conn.close()


@app.put("/api/v1/admin/lookups/{table_name}/values/{value_id}")
async def admin_update_lookup_value(table_name: str, value_id: int, payload: LookupValueUpdate, request: Request):
    p = await require_permission(request, "manage_system_settings")
    conn = await get_conn()
    try:
        t = await conn.fetchrow(
            "SELECT table_id FROM Lookup_Tables WHERE name = $1", table_name
        )
        if not t:
            raise HTTPException(404, f"Lookup table '{table_name}' not found")
        sets, vals = [], []
        if payload.value is not None:
            sets.append(f"value = ${len(vals)+1}"); vals.append(payload.value)
        if payload.label is not None:
            sets.append(f"label = ${len(vals)+1}"); vals.append(payload.label)
        if payload.display_order is not None:
            sets.append(f"display_order = ${len(vals)+1}"); vals.append(payload.display_order)
        if payload.is_active is not None:
            sets.append(f"is_active = ${len(vals)+1}"); vals.append(payload.is_active)
        if not sets:
            raise HTTPException(400, "No fields provided")
        vals.extend([value_id, t["table_id"]])
        row = await conn.fetchrow(
            f"UPDATE Lookup_Values SET {', '.join(sets)} WHERE value_id = ${len(vals)-1} AND table_id = ${len(vals)} RETURNING value_id, value, label, display_order, is_active",
            *vals,
        )
        if not row:
            raise HTTPException(404, "Value not found")
        return {"status": "ok", "value": dict(row)}
    finally:
        await conn.close()


@app.delete("/api/v1/admin/lookups/{table_name}/values/{value_id}")
async def admin_delete_lookup_value(table_name: str, value_id: int, request: Request):
    p = await require_permission(request, "manage_system_settings")
    conn = await get_conn()
    try:
        t = await conn.fetchrow(
            "SELECT table_id FROM Lookup_Tables WHERE name = $1", table_name
        )
        if not t:
            raise HTTPException(404, f"Lookup table '{table_name}' not found")
        result = await conn.execute(
            "DELETE FROM Lookup_Values WHERE value_id = $1 AND table_id = $2",
            value_id, t["table_id"],
        )
        if result == "DELETE 0":
            raise HTTPException(404, "Value not found")
        return {"status": "ok", "deleted": value_id}
    finally:
        await conn.close()


@app.post("/api/v1/admin/lookups/{table_name}/import")
async def admin_import_lookup_values(table_name: str, request: Request, file: UploadFile = File(...)):
    """Upload an Excel file to populate/update a lookup table.
    Expected format: single column with header 'value', optional 'label', 'display_order'.
    Existing values are preserved; new ones are added."""
    p = await require_permission(request, "manage_system_settings")
    conn = await get_conn()
    try:
        t = await conn.fetchrow(
            "SELECT table_id FROM Lookup_Tables WHERE name = $1", table_name
        )
        if not t:
            raise HTTPException(404, f"Lookup table '{table_name}' not found")

        contents = await file.read()
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        headers = [str(cell.value).strip().lower() for cell in ws[1]]
        if "value" not in headers:
            raise HTTPException(400, "Excel file must have a 'value' column header in row 1")

        vi = headers.index("value")
        li = headers.index("label") if "label" in headers else None
        di = headers.index("display_order") if "display_order" in headers else None

        added, skipped = 0, 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            val = str(row[vi]).strip() if row[vi] is not None else None
            if not val:
                skipped += 1
                continue
            label = str(row[li]).strip() if li is not None and row[li] is not None else val
            order = int(row[di]) if di is not None and row[di] is not None else row_idx
            try:
                await conn.execute(
                    """INSERT INTO Lookup_Values (table_id, value, label, display_order)
                       VALUES ($1, $2, $3, $4) ON CONFLICT (table_id, value) DO NOTHING""",
                    t["table_id"], val, label, order,
                )
                added += 1
            except Exception:
                skipped += 1

        return {"status": "ok", "added": added, "skipped": skipped, "table": table_name}
    finally:
        await conn.close()


@app.get("/api/v1/admin/lookups/{table_name}/export")
async def admin_export_lookup_values(table_name: str, request: Request):
    """Export a lookup table to an Excel file."""
    p = await require_permission(request, "manage_system_settings")
    conn = await get_conn()
    try:
        t = await conn.fetchrow(
            "SELECT table_id FROM Lookup_Tables WHERE name = $1", table_name
        )
        if not t:
            raise HTTPException(404, f"Lookup table '{table_name}' not found")
        values = await conn.fetch(
            """SELECT value, label, display_order, is_active
                 FROM Lookup_Values WHERE table_id = $1
              ORDER BY display_order ASC""",
            t["table_id"],
        )
    finally:
        await conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = table_name
    ws.append(["value", "label", "display_order", "is_active"])
    header_fill = PatternFill("solid", fgColor="00A651")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, 5):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
    for v in values:
        ws.append([v["value"], v["label"] or v["value"], v["display_order"], v["is_active"]])
    for col_letter in ["A", "B", "C", "D"]:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in ws[col_letter])
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{table_name}.xlsx"'},
    )


# ============================================================
#  Application endpoints
# ============================================================
def _validate_upload(upload):
    if upload.content_type:
        ok = (upload.content_type in ALLOWED_MIME_EXACT or
              any(upload.content_type.startswith(p) for p in ALLOWED_MIME_PREFIXES))
        if not ok:
            raise HTTPException(415, f"Unsupported file type: {upload.content_type}")


async def _save_upload(upload, app_id, attachment_type):
    _validate_upload(upload)
    contents = await upload.read()
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"File '{upload.filename}' exceeds {MAX_UPLOAD_BYTES // (1024*1024)}MB limit")
    original_name = Path(upload.filename or "file").name
    ext = Path(original_name).suffix.lower()
    stored_name = f"{attachment_type}_{uuid.uuid4().hex[:12]}{ext}"
    rel_path = f"app_{app_id}/{stored_name}"

    if USE_S3 and s3_client:
        s3_client.put_object(
            Bucket=MINIO_BUCKET,
            Key=rel_path,
            Body=contents,
            ContentType=upload.content_type or "application/octet-stream",
        )
    else:
        dest_path = UPLOADS_DIR / rel_path
        dest_path.parent.mkdir(exist_ok=True)
        with open(dest_path, "wb") as f:
            f.write(contents)

    return {
        "file_path": rel_path,
        "original_filename": original_name,
        "mime_type": upload.content_type,
        "file_size_bytes": len(contents),
    }


@app.post("/api/v1/applications/submit")
async def submit_application(
    request: Request,
    full_name: str  = Form(...),
    organization: str  = Form(...),
    email: Optional[str] = Form(None),
    form_variant: str = Form("full"),
    job_title: Optional[str] = Form(None),
    employee_number: Optional[str] = Form(None),
    subcontractor: Optional[str] = Form(None),
    gov_id_type: Optional[str] = Form(None),
    gov_id_number: Optional[str] = Form(None),
    blood_type: Optional[str] = Form(None),
    mobile_number: Optional[str] = Form(None),
    whatsapp_number: Optional[str] = Form(None),
    is_third_party_sponsor: bool = Form(False),
    date_joined: Optional[str] = Form(None),
    reports_to: Optional[str] = Form(None),
    iqama_expiry_date: Optional[str] = Form(None),
    passport_expiry_date: Optional[str] = Form(None),
    insurance_expiry_date: Optional[str] = Form(None),
    gender: Optional[str] = Form(None),
    nationality: Optional[str] = Form(None),
    date_of_birth: Optional[str] = Form(None),
    passport_number: Optional[str] = Form(None),
    language: Optional[str] = Form(None),
    safety_induction_date: Optional[str] = Form(None),
    sponsor_name: Optional[str] = Form(None),
    id_copy: Optional[UploadFile] = File(None),
    insurance: Optional[UploadFile] = File(None),
    photo: Optional[UploadFile] = File(None),
    legal_agreement: Optional[UploadFile] = File(None),
):
    p = await require_login(request)
    if p.must_change_password:
        raise HTTPException(409, "Password change required before continuing")

    submitter_user_id: Optional[int] = None
    submitter_external_id: Optional[int] = None
    if p.is_staff:
        # Staff need explicit submit_application permission
        conn0 = await get_conn()
        try:
            perms = await fetch_user_permissions(conn0, p.principal_id)
        finally:
            await conn0.close()
        if "submit_application" not in perms:
            raise HTTPException(403, "User lacks permission 'submit_application'")
        submitter_user_id = p.principal_id
    else:
        # External submitter: organization is locked to theirs; ignore the
        # `organization` form field and use their configured org instead.
        org_locked = p.extras.get("organization_name")
        if not org_locked:
            raise HTTPException(500, "External submitter has no associated organization")
        organization = org_locked
        # form_variant is also locked: group_2 -> 'full', group_1 -> 'brief'
        form_variant = "full" if p.extras.get("submitter_group") == "group_2" else "brief"
        submitter_external_id = p.principal_id

    conn = await get_conn()
    try:
        mode = await get_workflow_mode(conn)
        starting_state = "Pending HR Verification" if mode == "standard" else "Pending Security Clearance"
        date_joined_val = parse_date_or_none(date_joined)

        async with conn.transaction():
            iqama_expiry_val = parse_date_or_none(iqama_expiry_date)
            passport_expiry_val = parse_date_or_none(passport_expiry_date)
            insurance_expiry_val = parse_date_or_none(insurance_expiry_date)
            date_of_birth_val = parse_date_or_none(date_of_birth)
            safety_induction_val = parse_date_or_none(safety_induction_date)

            row = await conn.fetchrow(
                """INSERT INTO Site_Access_ID
                      (full_name, job_title, employee_number, organization,
                       subcontractor, gov_id_type, gov_id_number, blood_type,
                       mobile_number, whatsapp_number, email,
                       is_third_party_sponsor, submitted_by_user_id, submitted_by_external_id,
                       workflow_state, workflow_type,
                       form_variant, date_joined, reports_to,
                       iqama_expiry_date, passport_expiry_date, insurance_expiry_date,
                       gender, nationality, date_of_birth, passport_number,
                       language, safety_induction_date, sponsor_name)
                   VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18,$19,
                           $20,$21,$22,$23,$24,$25,$26,$27,$28,$29)
                RETURNING application_id, workflow_state""",
                full_name, job_title, employee_number, organization,
                subcontractor, gov_id_type, gov_id_number, blood_type,
                mobile_number, whatsapp_number, email,
                is_third_party_sponsor, submitter_user_id, submitter_external_id,
                starting_state, mode,
                form_variant, date_joined_val, reports_to,
                iqama_expiry_val, passport_expiry_val, insurance_expiry_val,
                gender, nationality, date_of_birth_val, passport_number,
                language, safety_induction_val, sponsor_name,
            )
            app_id = row["application_id"]

            uploads = {"id_copy": id_copy, "insurance": insurance, "photo": photo, "legal_agreement": legal_agreement}
            for kind, upload in uploads.items():
                if upload is None or upload.filename in (None, ""):
                    continue
                meta = await _save_upload(upload, app_id, kind)
                await conn.execute(
                    """INSERT INTO Attachments
                          (application_id, attachment_type, file_path, original_filename,
                           mime_type, file_size_bytes, uploaded_by_user_id)
                       VALUES ($1,$2,$3,$4,$5,$6,$7)""",
                    app_id, kind, meta["file_path"], meta["original_filename"],
                    meta["mime_type"], meta["file_size_bytes"], submitter_user_id,
                )

            history_comment = (
                f"Application submitted ({form_variant} form) "
                f"by {'staff' if p.is_staff else 'external submitter'}: {p.full_name}"
            )
            await conn.execute(
                """INSERT INTO Workflow_History
                      (application_id, action_taken, previous_state, new_state, acted_by_user_id, comments)
                   VALUES ($1, 'Submitted', 'Draft', $2, $3, $4)""",
                app_id, row["workflow_state"], submitter_user_id, history_comment,
            )
        return {"status": "ok", "application_id": app_id}
    finally:
        await conn.close()


@app.get("/api/v1/applications/pending")
async def list_pending(request: Request):
    await require_permission(request, "view_all_applications")
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT application_id, full_name, organization, workflow_state, created_at
                 FROM Site_Access_ID
                WHERE workflow_state NOT IN ('Approved','Rejected','Cancelled')
             ORDER BY application_id DESC"""
        )
        return {"pending_applications": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.get("/api/v1/applications")
async def list_all_applications(request: Request):
    await require_permission(request, "view_all_applications")
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT application_id, full_name, organization, workflow_state, created_at
                 FROM Site_Access_ID
             ORDER BY application_id DESC"""
        )
        return {"applications": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.get("/api/v1/applications/export")
async def export_applications_xlsx(request: Request):
    await require_permission(request, "view_all_applications")
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT s.application_id, s.created_at, s.workflow_state, s.workflow_type, s.form_variant,
                      s.full_name, s.email, s.mobile_number, s.whatsapp_number,
                      s.gov_id_type, s.gov_id_number, s.blood_type,
                      s.job_title, s.employee_number, s.organization, s.subcontractor,
                      s.is_third_party_sponsor, s.date_joined, s.reports_to,
                      s.iqama_expiry_date, s.passport_expiry_date, s.insurance_expiry_date,
                      s.gender, s.nationality, s.date_of_birth, s.passport_number,
                      s.language, s.safety_induction_date, s.sponsor_name,
                      u.full_name AS submitter_name
                 FROM Site_Access_ID s
            LEFT JOIN App_Users u ON u.user_id = s.submitted_by_user_id
             ORDER BY s.application_id ASC"""
        )
    finally:
        await conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    headers = [
        "App ID", "Submitted At", "Status", "Workflow Type", "Form Variant",
        "Full Name", "Email", "Mobile", "WhatsApp",
        "Gov ID Type", "Gov ID Number", "Blood Type",
        "Job/Trade", "Employee #", "Organization", "Subcontractor",
        "3rd-Party Sponsor", "Date Joined", "Reports To",
        "Iqama Expiry", "Passport Expiry", "Insurance Expiry",
        "Gender", "Nationality", "Date of Birth", "Passport Number",
        "Language", "Safety Induction Date", "Sponsor Name",
        "Submitted By",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="00A651")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append([
            r["application_id"],
            r["created_at"].strftime("%Y-%m-%d %H:%M") if r["created_at"] else "",
            r["workflow_state"], r["workflow_type"], r["form_variant"] or "",
            r["full_name"], r["email"] or "", r["mobile_number"] or "", r["whatsapp_number"] or "",
            r["gov_id_type"] or "", r["gov_id_number"] or "", r["blood_type"] or "",
            r["job_title"] or "", r["employee_number"] or "",
            r["organization"] or "", r["subcontractor"] or "",
            "Yes" if r["is_third_party_sponsor"] else "No",
            r["date_joined"].strftime("%Y-%m-%d") if r["date_joined"] else "",
            r["reports_to"] or "",
            r["iqama_expiry_date"].strftime("%Y-%m-%d") if r["iqama_expiry_date"] else "",
            r["passport_expiry_date"].strftime("%Y-%m-%d") if r["passport_expiry_date"] else "",
            r["insurance_expiry_date"].strftime("%Y-%m-%d") if r["insurance_expiry_date"] else "",
            r["gender"] or "", r["nationality"] or "",
            r["date_of_birth"].strftime("%Y-%m-%d") if r["date_of_birth"] else "",
            r["passport_number"] or "",
            r["language"] or "",
            r["safety_induction_date"].strftime("%Y-%m-%d") if r["safety_induction_date"] else "",
            r["sponsor_name"] or "",
            r["submitter_name"] or "",
        ])
    for i in range(len(headers)):
        col_letter = get_column_letter(i + 1)
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in ws[col_letter])
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"taawon_p3_applications_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# New column name normalisation helper for Excel imports
_IMPORT_ALIASES = {
    "full_name": ("full_name", "fullname", "full name", "name", "applicant name"),
    "organization": ("organization", "company", "company name", "employer", "org"),
    "subcontractor": ("subcontractor", "sub contractor", "sub"),
    "gov_id_type": ("gov_id_type", "gov id type", "id type", "gov_id"),
    "gov_id_number": ("gov_id_number", "gov id number", "id number", "iqama number", "national id", "gov_id"),
    "blood_type": ("blood_type", "bloodtype", "blood type"),
    "mobile_number": ("mobile_number", "mobile number", "mobile", "phone", "phone number", "contact"),
    "whatsapp_number": ("whatsapp_number", "whatsapp number", "whatsapp"),
    "email": ("email", "e-mail", "mail"),
    "employee_number": ("employee_number", "employee number", "employee no", "emp no", "emp_number", "empno"),
    "date_joined": ("date_joined", "date joined", "joined date", "join date"),
    "reports_to": ("reports_to", "reports to", "report to", "supervisor"),
    "full_name_arabic": ("full_name_arabic", "full name arabic", "arabic name", "name_ar", "arabic"),
}

def _normalise_import_header(raw: str) -> str:
    """Map an Excel column header to the canonical field name."""
    raw = raw.strip().lower().replace("_", " ").replace("-", " ")
    for canonical, aliases in _IMPORT_ALIASES.items():
        if raw in aliases or raw.replace(" ", "_") in aliases:
            return canonical
    return raw  # pass through unknown headers


@app.post("/api/v1/admin/applications/import")
async def admin_import_applications(request: Request, file: UploadFile = File(...)):
    """Bulk-import backlog applications from an Excel file.

    Expected columns (first row = headers; flexible naming accepted):
      full_name*, organization*, employee_number, gov_id_number,
      blood_type, mobile_number, whatsapp_number, email, date_joined,
      reports_to

    * required. Rows missing either are skipped.
    New records get workflow_state = 'Approved' and a history entry.
    """
    await require_permission(request, "manage_system_settings")
    contents = await file.read()
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(contents))
    ws = wb.active
    headers = [_normalise_import_header(str(cell.value or "")) for cell in ws[1]]

    required = ("full_name", "organization")
    missing = [c for c in required if c not in headers]
    if missing:
        raise HTTPException(400, f"Missing required column(s): {', '.join(missing)}")

    fi = {h: i for i, h in enumerate(headers)}

    def _val(row, field):
        idx = fi.get(field)
        if idx is None:
            return None
        v = row[idx]
        if v is None:
            return None
        return str(v).strip()

    conn = await get_conn()
    try:
        imported, skipped, errors = 0, 0, []
        app_ids = []
        mode = "security_only"

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            full_name = _val(row, "full_name")
            organization = _val(row, "organization")
            if not full_name or not organization:
                skipped += 1
                errors.append(f"Row {row_idx}: missing required field(s) — skipped")
                continue

            try:
                async with conn.transaction():
                    r = await conn.fetchrow(
                        """INSERT INTO Site_Access_ID
                              (full_name, organization, subcontractor,
                               gov_id_type, gov_id_number, blood_type,
                               mobile_number, whatsapp_number, email,
                               employee_number, date_joined, reports_to,
                               workflow_state, workflow_type, form_variant,
                               submitted_by_user_id)
                           VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,
                                   'Approved', $13, 'brief', $14)
                           RETURNING application_id""",
                        full_name,
                        organization,
                        _val(row, "subcontractor"),
                        _val(row, "gov_id_type") or "Iqama",
                        _val(row, "gov_id_number"),
                        _val(row, "blood_type"),
                        _val(row, "mobile_number"),
                        _val(row, "whatsapp_number"),
                        _val(row, "email"),
                        _val(row, "employee_number"),
                        parse_date_or_none(_val(row, "date_joined")),
                        _val(row, "reports_to"),
                        mode,
                        None,  # imported by system, not a specific user
                    )
                    app_id = r["application_id"]
                    await conn.execute(
                        """INSERT INTO Workflow_History
                              (application_id, action_taken, previous_state, new_state,
                               acted_by_user_id, comments)
                           VALUES ($1, 'Bulk Import', 'Draft', 'Approved', $2,
                                   'Imported via bulk backlog upload')""",
                        app_id, None,
                    )
                    imported += 1
                    app_ids.append(app_id)
            except Exception as exc:
                skipped += 1
                errors.append(f"Row {row_idx} ({full_name}): {exc}")

        return {
            "status": "ok",
            "imported": imported,
            "skipped": skipped,
            "errors": errors[:50],  # cap to avoid giant response
            "application_ids": app_ids,
        }
    finally:
        await conn.close()


@app.get("/api/v1/admin/non-grata")
async def list_non_grata(request: Request):
    await require_permission(request, "manage_system_settings")
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            "SELECT ng_id, full_name, gender, nationality, gov_id_type, gov_id_number, "
            "date_of_birth, created_at FROM Non_Grata ORDER BY ng_id DESC"
        )
        return {"entries": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.post("/api/v1/admin/non-grata/import")
async def import_non_grata(request: Request, file: UploadFile = File(...)):
    await require_permission(request, "manage_system_settings")
    conn = await get_conn()
    try:
        contents = await file.read()
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

        aliases = {
            "full_name": ("full_name", "fullname", "full name", "name"),
            "gender": ("gender", "sex"),
            "nationality": ("nationality", "nation"),
            "gov_id_type": ("gov_id_type", "gov id type", "id type", "id_type"),
            "gov_id_number": ("gov_id_number", "gov id number", "id number", "id_number"),
            "date_of_birth": ("date_of_birth", "date of birth", "dob", "birth date"),
        }

        col = {}
        for key, alts in aliases.items():
            for h in headers:
                if h in alts:
                    col[key] = headers.index(h)
                    break

        if "full_name" not in col:
            raise HTTPException(400, "Excel must have a 'full_name' or 'name' column")

        imported, skipped = 0, 0
        errors = []
        async with conn.transaction():
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    full_name = str(row[col["full_name"]]).strip() if row[col["full_name"]] else ""
                    if not full_name:
                        skipped += 1
                        continue
                    gender = str(row[col["gender"]]).strip() if "gender" in col and row[col["gender"]] else None
                    nationality = str(row[col["nationality"]]).strip() if "nationality" in col and row[col["nationality"]] else None
                    gov_id_type = str(row[col["gov_id_type"]]).strip() if "gov_id_type" in col and row[col["gov_id_type"]] else None
                    gov_id_number = str(row[col["gov_id_number"]]).strip() if "gov_id_number" in col and row[col["gov_id_number"]] else None
                    dob_str = str(row[col["date_of_birth"]]).strip() if "date_of_birth" in col and row[col["date_of_birth"]] else None
                    dob = parse_date_or_none(dob_str)
                    await conn.execute(
                        """INSERT INTO Non_Grata
                              (full_name, gender, nationality, gov_id_type, gov_id_number, date_of_birth)
                           VALUES ($1,$2,$3,$4,$5,$6)""",
                        full_name, gender, nationality, gov_id_type, gov_id_number, dob,
                    )
                    imported += 1
                except Exception as exc:
                    skipped += 1
                    errors.append(f"Row {row_idx}: {exc}")
        return {"status": "ok", "imported": imported, "skipped": skipped, "errors": errors[:50]}
    finally:
        await conn.close()


@app.get("/api/v1/applications/{application_id}")
async def get_application_full(application_id: int, request: Request):
    await require_permission(request, "view_all_applications")
    conn = await get_conn()
    try:
        app_row = await conn.fetchrow(
            """SELECT s.*, u.full_name AS submitter_name
                 FROM Site_Access_ID s
            LEFT JOIN App_Users u ON u.user_id = s.submitted_by_user_id
                WHERE s.application_id = $1""",
            application_id,
        )
        if not app_row:
            raise HTTPException(404, "Application not found")
        history_rows = await conn.fetch(
            """SELECT h.*, u.full_name AS acted_by
                 FROM Workflow_History h
            LEFT JOIN App_Users u ON u.user_id = h.acted_by_user_id
                WHERE h.application_id = $1
             ORDER BY h.action_timestamp ASC""",
            application_id,
        )
        attach_rows = await conn.fetch(
            """SELECT attachment_id, attachment_type, original_filename,
                      mime_type, file_size_bytes, uploaded_at
                 FROM Attachments
                WHERE application_id = $1
             ORDER BY attachment_id ASC""",
            application_id,
        )
        notif_rows = await conn.fetch(
            """SELECT notification_id, channel, recipient, subject, body,
                      delivery_status, error_message, created_at, sent_at
                 FROM Notifications
                WHERE application_id = $1
             ORDER BY notification_id DESC""",
            application_id,
        )
        return {
            "application":   dict(app_row),
            "history":       [dict(r) for r in history_rows],
            "attachments":   [dict(r) for r in attach_rows],
            "notifications": [dict(r) for r in notif_rows],
        }
    finally:
        await conn.close()


@app.put("/api/v1/applications/{application_id}")
async def edit_application_fields(application_id: int, payload: ApplicationFieldsUpdate, request: Request):
    p = await require_permission(request, "edit_application")
    conn = await get_conn()
    try:
        app_row = await conn.fetchrow(
            "SELECT workflow_state FROM Site_Access_ID WHERE application_id = $1",
            application_id,
        )
        if not app_row:
            raise HTTPException(404, "Application not found")
        if app_row["workflow_state"] in ("Approved", "Rejected", "Cancelled"):
            raise HTTPException(409, f"Application is in '{app_row['workflow_state']}' state and cannot be edited")
        editable_fields = [
            "full_name", "job_title", "employee_number", "organization", "subcontractor",
            "gov_id_type", "gov_id_number", "blood_type", "mobile_number", "whatsapp_number",
            "email", "is_third_party_sponsor", "reports_to",
            "gender", "nationality", "date_of_birth", "passport_number",
            "language", "safety_induction_date", "sponsor_name",
        ]
        sets, vals = [], []
        changed_summary = []
        for f in editable_fields:
            v = getattr(payload, f)
            if v is not None:
                sets.append(f"{f} = ${len(vals) + 1}")
                vals.append(v)
                changed_summary.append(f)
        date_fields = ["date_joined", "iqama_expiry_date", "passport_expiry_date", "insurance_expiry_date"]
        for f in date_fields:
            v = getattr(payload, f)
            if v is not None:
                sets.append(f"{f} = ${len(vals) + 1}")
                vals.append(parse_date_or_none(v))
                changed_summary.append(f)
        if not sets:
            raise HTTPException(400, "No editable fields provided")
        vals.append(application_id)
        async with conn.transaction():
            await conn.execute(
                f"UPDATE Site_Access_ID SET {', '.join(sets)} WHERE application_id = ${len(vals)}",
                *vals,
            )
            await conn.execute(
                """INSERT INTO Workflow_History
                      (application_id, action_taken, previous_state, new_state,
                       acted_by_user_id, comments)
                   VALUES ($1, 'Edit', $2, $2, $3, $4)""",
                application_id, app_row["workflow_state"], p.principal_id,
                f"Fields edited: {', '.join(changed_summary)}",
            )
        return {"status": "ok", "fields_changed": changed_summary}
    finally:
        await conn.close()


@app.post("/api/v1/applications/{application_id}/attachments")
async def add_application_attachment(
    application_id: int, request: Request,
    attachment_type: str = Form(...),
    file: UploadFile = File(...),
):
    """Upload an attachment to an existing (pending) application."""
    p = await require_permission(request, "edit_application")
    conn = await get_conn()
    try:
        app_row = await conn.fetchrow(
            "SELECT workflow_state FROM Site_Access_ID WHERE application_id = $1",
            application_id,
        )
        if not app_row:
            raise HTTPException(404, "Application not found")
        if app_row["workflow_state"] in ("Approved", "Rejected", "Cancelled"):
            raise HTTPException(409, f"Application is in '{app_row['workflow_state']}' state — cannot modify attachments")
    finally:
        await conn.close()

    meta = await _save_upload(file, application_id, attachment_type)

    conn = await get_conn()
    try:
        await conn.execute(
            """INSERT INTO Attachments
                  (application_id, attachment_type, file_path, original_filename,
                   mime_type, file_size_bytes, uploaded_by_user_id)
               VALUES ($1,$2,$3,$4,$5,$6,$7)""",
            application_id, attachment_type, meta["file_path"],
            meta["original_filename"], meta["mime_type"],
            meta["file_size_bytes"], None,
        )
        await conn.execute(
            """INSERT INTO Workflow_History
                  (application_id, action_taken, previous_state, new_state,
                   acted_by_user_id, comments)
               VALUES ($1, 'Edit', $2, $2, $3, $4)""",
            application_id, app_row["workflow_state"], p.principal_id,
            f"Attachment added: {attachment_type} ({meta['original_filename']})",
        )
        return {"status": "ok", "file_path": meta["file_path"], "original_filename": meta["original_filename"]}
    finally:
        await conn.close()


@app.delete("/api/v1/applications/{application_id}/attachments/{attachment_id}")
async def delete_application_attachment(
    application_id: int, attachment_id: int, request: Request,
):
    """Delete an attachment from a pending application."""
    p = await require_permission(request, "edit_application")
    conn = await get_conn()
    try:
        app_row = await conn.fetchrow(
            "SELECT workflow_state FROM Site_Access_ID WHERE application_id = $1",
            application_id,
        )
        if not app_row:
            raise HTTPException(404, "Application not found")
        if app_row["workflow_state"] in ("Approved", "Rejected", "Cancelled"):
            raise HTTPException(409, f"Application is in '{app_row['workflow_state']}' state — cannot modify attachments")

        att = await conn.fetchrow(
            "SELECT attachment_id, attachment_type, file_path, original_filename FROM Attachments WHERE attachment_id = $1 AND application_id = $2",
            attachment_id, application_id,
        )
        if not att:
            raise HTTPException(404, "Attachment not found")

        # Delete the file from storage
        if USE_S3 and s3_client:
            try:
                s3_client.delete_object(Bucket=MINIO_BUCKET, Key=att["file_path"])
            except Exception:
                pass  # best-effort
        else:
            full_path = (UPLOADS_DIR / att["file_path"]).resolve()
            try:
                full_path.relative_to(UPLOADS_DIR.resolve())
                if full_path.is_file():
                    full_path.unlink()
            except ValueError:
                pass

        await conn.execute(
            "DELETE FROM Attachments WHERE attachment_id = $1", attachment_id,
        )
        await conn.execute(
            """INSERT INTO Workflow_History
                  (application_id, action_taken, previous_state, new_state,
                   acted_by_user_id, comments)
               VALUES ($1, 'Edit', $2, $2, $3, $4)""",
            application_id, app_row["workflow_state"], p.principal_id,
            f"Attachment removed: {att['attachment_type']} ({att['original_filename']})",
        )
        return {"status": "ok", "deleted_attachment_id": attachment_id}
    finally:
        await conn.close()


@app.get("/api/v1/applications/{application_id}/attachment/{attachment_id}")
async def download_attachment(
    application_id: int, attachment_id: int, request: Request,
    preview: bool = False,
):
    p = await require_login(request)
    if p.must_change_password:
        raise HTTPException(409, "Password change required before continuing")
    conn = await get_conn()
    try:
        # Auth check: staff with view_all_applications, OR the original submitter (staff/external)
        if p.is_staff:
            perms = await fetch_user_permissions(conn, p.principal_id)
            allowed = "view_all_applications" in perms
            if not allowed:
                owner = await conn.fetchrow(
                    "SELECT submitted_by_user_id FROM Site_Access_ID WHERE application_id = $1",
                    application_id,
                )
                allowed = owner and owner["submitted_by_user_id"] == p.principal_id
        else:
            owner = await conn.fetchrow(
                "SELECT submitted_by_external_id FROM Site_Access_ID WHERE application_id = $1",
                application_id,
            )
            allowed = owner and owner["submitted_by_external_id"] == p.principal_id
        if not allowed:
            raise HTTPException(403, "Not allowed to access this attachment")

        row = await conn.fetchrow(
            """SELECT file_path, original_filename, mime_type
                 FROM Attachments
                WHERE attachment_id = $1 AND application_id = $2""",
            attachment_id, application_id,
        )
    finally:
        await conn.close()
    if not row:
        raise HTTPException(404, "Attachment not found")

    file_path = row["file_path"]
    media_type = row["mime_type"] or "application/octet-stream"
    filename = row["original_filename"]
    disposition = f'inline; filename="{filename}"' if preview else f'attachment; filename="{filename}"'

    if USE_S3 and s3_client:
        try:
            obj = s3_client.get_object(Bucket=MINIO_BUCKET, Key=file_path)
            content = obj["Body"].read()
        except Exception:
            raise HTTPException(410, "File missing on storage")
        return Response(content=content, media_type=media_type,
                        headers={"Content-Disposition": disposition})

    full_path = (UPLOADS_DIR / file_path).resolve()
    try:
        full_path.relative_to(UPLOADS_DIR.resolve())
    except ValueError:
        raise HTTPException(400, "Invalid path")
    if not full_path.is_file():
        raise HTTPException(410, "File missing on disk")
    if preview:
        return FileResponse(
            full_path,
            media_type=media_type,
            headers={"Content-Disposition": disposition},
        )
    return FileResponse(
        full_path,
        media_type=media_type,
        filename=filename,
    )


@app.put("/api/v1/applications/{application_id}/status")
async def update_status(application_id: int, payload: StatusUpdate, request: Request):
    p = await require_staff(request)
    conn = await get_conn()
    try:
        async with conn.transaction():
            app_row = await conn.fetchrow(
                "SELECT * FROM Site_Access_ID WHERE application_id = $1",
                application_id,
            )
            if not app_row:
                raise HTTPException(404, "Application not found")
            current_state = app_row["workflow_state"]
            required = await required_permission_for_state_change(conn, current_state, payload.new_state)
            if required:
                perms = await fetch_user_permissions(conn, p.principal_id)
                if required not in perms:
                    raise HTTPException(
                        403,
                        f"User lacks permission '{required}' (transition: {current_state} -> {payload.new_state})",
                    )
            await conn.execute(
                "UPDATE Site_Access_ID SET workflow_state = $1 WHERE application_id = $2",
                payload.new_state, application_id,
            )
            await conn.execute(
                """INSERT INTO Workflow_History
                      (application_id, action_taken, previous_state, new_state,
                       acted_by_user_id, comments)
                   VALUES ($1, $2, $3, $4, $5, $6)""",
                application_id,
                "Reject" if payload.new_state == "Rejected" else "Approve",
                current_state, payload.new_state,
                p.principal_id, payload.comments,
            )
            if payload.new_state == "Rejected":
                reason = payload.comments or "(no reason provided)"
                await _handle_rejection_notifications(conn, app_row, reason)
        return {"status": "ok", "application_id": application_id, "new_state": payload.new_state}
    finally:
        await conn.close()


@app.get("/api/v1/history")
async def list_history(request: Request):
    await require_permission(request, "view_all_applications")
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT h.*, u.full_name AS acted_by
                 FROM Workflow_History h
            LEFT JOIN App_Users u ON u.user_id = h.acted_by_user_id
             ORDER BY h.action_timestamp DESC
                LIMIT 500"""
        )
        return {"history": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.get("/api/v1/admin/induction-workbench")
async def list_induction_workbench(request: Request):
    await require_permission(request, "manage_system_settings")
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            "SELECT iw_id, full_name, gender, nationality, date_of_birth, gov_id_number, "
            "mobile_number, blood_type, language, company, legal_agreement, insurance_validity, "
            "profession_iqama, profession_work, location, induction_date, status, created_at "
            "FROM Induction_Workbench ORDER BY iw_id DESC"
        )
        return {"entries": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.post("/api/v1/admin/induction-workbench/import")
async def import_induction_workbench(request: Request, file: UploadFile = File(...)):
    await require_permission(request, "manage_system_settings")
    conn = await get_conn()
    try:
        contents = await file.read()
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

        aliases = {
            "full_name": ("full_name", "fullname", "full name", "name", "employee name"),
            "gender": ("gender", "sex"),
            "nationality": ("nationality", "nation"),
            "date_of_birth": ("date_of_birth", "date of birth", "dob", "birth date"),
            "gov_id_number": ("gov_id_number", "gov id number", "id number", "id_number", "id/ iq #", "id/iq #", "iqama/id"),
            "mobile_number": ("mobile_number", "mobile number", "mobil #", "mobil", "mobile", "phone"),
            "blood_type": ("blood_type", "blood type", "blood"),
            "language": ("language", "lang"),
            "company": ("company", "org", "organization"),
            "legal_agreement": ("legal_agreement", "legal agreement", "agreement"),
            "insurance_validity": ("insurance_validity", "insurance validity", "insurance", "insurance valid"),
            "profession_iqama": ("profession_iqama", "profession iqama", "iqama profession", "profession"),
            "profession_work": ("profession_work", "profession work", "work profession", "job title"),
            "location": ("location", "loc"),
            "induction_date": ("induction_date", "induction date", "induction"),
            "status": ("status", "state"),
        }

        col = {}
        for key, alts in aliases.items():
            for h in headers:
                if h in alts:
                    col[key] = headers.index(h)
                    break

        if "full_name" not in col:
            raise HTTPException(400, "Excel must have a 'full_name' or 'name' column")

        imported, skipped = 0, 0
        errors = []
        async with conn.transaction():
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    full_name = str(row[col["full_name"]]).strip() if row[col["full_name"]] else ""
                    if not full_name:
                        skipped += 1
                        continue
                    gender = str(row[col["gender"]]).strip() if "gender" in col and row[col["gender"]] else None
                    nationality = str(row[col["nationality"]]).strip() if "nationality" in col and row[col["nationality"]] else None
                    dob_str = str(row[col["date_of_birth"]]).strip() if "date_of_birth" in col and row[col["date_of_birth"]] else None
                    dob = parse_date_or_none(dob_str)
                    gov_id_number = str(row[col["gov_id_number"]]).strip() if "gov_id_number" in col and row[col["gov_id_number"]] else None
                    mobile_number = str(row[col["mobile_number"]]).strip() if "mobile_number" in col and row[col["mobile_number"]] else None
                    blood_type = str(row[col["blood_type"]]).strip() if "blood_type" in col and row[col["blood_type"]] else None
                    language = str(row[col["language"]]).strip() if "language" in col and row[col["language"]] else None
                    company = str(row[col["company"]]).strip() if "company" in col and row[col["company"]] else None
                    legal_agreement = str(row[col["legal_agreement"]]).strip() if "legal_agreement" in col and row[col["legal_agreement"]] else None
                    ins_str = str(row[col["insurance_validity"]]).strip() if "insurance_validity" in col and row[col["insurance_validity"]] else None
                    insurance_validity = parse_date_or_none(ins_str)
                    profession_iqama = str(row[col["profession_iqama"]]).strip() if "profession_iqama" in col and row[col["profession_iqama"]] else None
                    profession_work = str(row[col["profession_work"]]).strip() if "profession_work" in col and row[col["profession_work"]] else None
                    location = str(row[col["location"]]).strip() if "location" in col and row[col["location"]] else None
                    ind_str = str(row[col["induction_date"]]).strip() if "induction_date" in col and row[col["induction_date"]] else None
                    induction_date = parse_date_or_none(ind_str)
                    status = str(row[col["status"]]).strip() if "status" in col and row[col["status"]] else None
                    await conn.execute(
                        """INSERT INTO Induction_Workbench
                              (full_name, gender, nationality, date_of_birth, gov_id_number,
                               mobile_number, blood_type, language, company, legal_agreement,
                               insurance_validity, profession_iqama, profession_work, location,
                               induction_date, status)
                           VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16)""",
                        full_name, gender, nationality, dob, gov_id_number,
                        mobile_number, blood_type, language, company, legal_agreement,
                        insurance_validity, profession_iqama, profession_work, location,
                        induction_date, status,
                    )
                    imported += 1
                except Exception as exc:
                    skipped += 1
                    errors.append(f"Row {row_idx}: {exc}")
        return {"status": "ok", "imported": imported, "skipped": skipped, "errors": errors[:50]}
    finally:
        await conn.close()


@app.get("/api/v1/induction-workbench/lookup")
async def lookup_induction_workbench(request: Request, q: str = ""):
    await require_permission(request, "view_all_applications")
    if not q or len(q.strip()) < 2:
        return {"matches": []}
    q = q.strip()
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT iw_id, full_name, gender, nationality, date_of_birth, gov_id_number,
                      mobile_number, blood_type, language, company, legal_agreement,
                      insurance_validity, profession_iqama, profession_work, location,
                      induction_date, status
                 FROM Induction_Workbench
                WHERE full_name ILIKE $1 OR gov_id_number ILIKE $1
                LIMIT 20""",
            f"%{q}%",
        )
        return {"matches": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.get("/api/v1/admin/system-users")
async def list_system_users(request: Request):
    await require_permission(request, "manage_system_settings")
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            "SELECT su_id, employee_name, gov_id_number, email, project, company, "
            "job_title, location, mobile_number, nationality, role, created_at "
            "FROM System_Users ORDER BY su_id DESC"
        )
        return {"entries": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.post("/api/v1/admin/system-users/import")
async def import_system_users(request: Request, file: UploadFile = File(...)):
    await require_permission(request, "manage_system_settings")
    conn = await get_conn()
    try:
        contents = await file.read()
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

        aliases = {
            "employee_name": ("employee_name", "employee name", "full name", "name"),
            "gov_id_number": ("gov_id_number", "gov id number", "id number", "iqama/id", "iqama / id", "id"),
            "email": ("email", "e-mail", "mail"),
            "project": ("project", "proj", "project name"),
            "company": ("company", "org", "organization"),
            "job_title": ("job_title", "job title", "title", "position", "role title"),
            "location": ("location", "loc", "site"),
            "mobile_number": ("mobile_number", "mobile number", "mobile #", "mobile", "phone"),
            "nationality": ("nationality", "nation"),
            "role": ("role", "user role", "system role"),
        }

        col = {}
        for key, alts in aliases.items():
            for h in headers:
                if h in alts:
                    col[key] = headers.index(h)
                    break

        if "employee_name" not in col:
            raise HTTPException(400, "Excel must have an 'employee_name' or 'name' column")

        imported, skipped = 0, 0
        errors = []
        async with conn.transaction():
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    employee_name = str(row[col["employee_name"]]).strip() if row[col["employee_name"]] else ""
                    if not employee_name:
                        skipped += 1
                        continue
                    gov_id_number = str(row[col["gov_id_number"]]).strip() if "gov_id_number" in col and row[col["gov_id_number"]] else None
                    email = str(row[col["email"]]).strip() if "email" in col and row[col["email"]] else None
                    project = str(row[col["project"]]).strip() if "project" in col and row[col["project"]] else None
                    company = str(row[col["company"]]).strip() if "company" in col and row[col["company"]] else None
                    job_title = str(row[col["job_title"]]).strip() if "job_title" in col and row[col["job_title"]] else None
                    location = str(row[col["location"]]).strip() if "location" in col and row[col["location"]] else None
                    mobile_number = str(row[col["mobile_number"]]).strip() if "mobile_number" in col and row[col["mobile_number"]] else None
                    nationality = str(row[col["nationality"]]).strip() if "nationality" in col and row[col["nationality"]] else None
                    role = str(row[col["role"]]).strip() if "role" in col and row[col["role"]] else None
                    await conn.execute(
                        """INSERT INTO System_Users
                              (employee_name, gov_id_number, email, project, company, job_title,
                               location, mobile_number, nationality, role)
                           VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10)""",
                        employee_name, gov_id_number, email, project, company, job_title,
                        location, mobile_number, nationality, role,
                    )
                    imported += 1
                except Exception as exc:
                    skipped += 1
                    errors.append(f"Row {row_idx}: {exc}")
        return {"status": "ok", "imported": imported, "skipped": skipped, "errors": errors[:50]}
    finally:
        await conn.close()


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    return JSONResponse({"detail": exc.detail}, status_code=exc.status_code)
