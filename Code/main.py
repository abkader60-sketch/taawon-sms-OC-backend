"""
Ta'awon P3 SMS - FastAPI Backend
================================
Single-file backend.

Changes in this revision (1 May 2026, v0.7 - "pre-deployment prep"):
  + DATABASE_URL support: when set (Railway auto-sets it), it overrides the
    individual DB_HOST/DB_PORT/DB_NAME/DB_USER/DB_PASSWORD vars. asyncpg
    accepts a full DSN directly.
  + Configurable CORS_ORIGINS via env var. Defaults to '*' for local dev.
    Production should set this to the actual frontend origin.
  + Migration fixed: Omar->Administrator now uses full_name lookup, not the
    brittle user_id=2 hardcode that bit us last session.
  + No longer ships with a hardcoded DB_PASSWORD fallback. If neither
    DATABASE_URL nor DB_PASSWORD is set, the app refuses to start.

NOT in this revision (still on the roadmap for v0.8 = real auth):
  - Real login (still using role-picker + X-Acting-User-Id header)
  - bcrypt-hashed staff passwords
  - Auth-gated attachment downloads
  - Rate limiting

Backwards-compatible additive migrations.
"""

import os
import secrets
import smtplib
import ssl
import sys
import uuid
from datetime import datetime, date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO
from pathlib import Path
from typing import List, Optional

import asyncpg
import bcrypt
from dotenv import load_dotenv
from fastapi import (FastAPI, File, Form, Header, HTTPException, Path as FPath,
                     Request, UploadFile)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel

# ============================================================
#  Configuration
# ============================================================
PROJECT_ROOT = Path(__file__).resolve().parent.parent
ENV_FILE = PROJECT_ROOT / ".env"
load_dotenv(ENV_FILE)
# Also try the same dir as main.py (Railway puts .env there if uploaded)
load_dotenv(Path(__file__).resolve().parent / ".env")

DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
DB_HOST     = os.getenv("DB_HOST", "localhost")
DB_PORT     = int(os.getenv("DB_PORT", "5432"))
DB_NAME     = os.getenv("DB_NAME", "id_system")
DB_USER     = os.getenv("DB_USER", "postgres")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")

if not DATABASE_URL and not DB_PASSWORD:
    print("ERROR: No database configuration found.", file=sys.stderr)
    print("Set either DATABASE_URL (preferred for production) or DB_PASSWORD.", file=sys.stderr)
    print("See .env.example for guidance.", file=sys.stderr)
    sys.exit(1)

SMTP_HOST     = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT     = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
SMTP_FROM     = os.getenv("SMTP_FROM", SMTP_USERNAME)
SMTP_ENABLED  = bool(SMTP_USERNAME and SMTP_PASSWORD)

# CORS - comma-separated list of origins. '*' allows all (dev only).
_cors_raw = os.getenv("CORS_ORIGINS", "*").strip()
CORS_ORIGINS = [o.strip() for o in _cors_raw.split(",") if o.strip()] or ["*"]

UPLOADS_DIR = Path(os.getenv("UPLOADS_DIR", str(PROJECT_ROOT / "uploads")))
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
ALLOWED_MIME_PREFIXES = ("image/",)
ALLOWED_MIME_EXACT = {"application/pdf"}

# ============================================================
#  App setup
# ============================================================
app = FastAPI(title="Ta'awon P3 SMS API", version="0.7.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


async def get_conn() -> asyncpg.Connection:
    """Get a fresh DB connection. Prefers DATABASE_URL if set."""
    if DATABASE_URL:
        return await asyncpg.connect(DATABASE_URL)
    return await asyncpg.connect(
        host=DB_HOST, port=DB_PORT, database=DB_NAME,
        user=DB_USER, password=DB_PASSWORD
    )


# ============================================================
#  Schema bootstrap (additive - safe to re-run)
# ============================================================
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS App_Users (
    user_id              SERIAL PRIMARY KEY,
    full_name            VARCHAR(100) NOT NULL,
    username             VARCHAR(50) UNIQUE NOT NULL,
    hashed_password      VARCHAR(255) NOT NULL,
    force_password_change BOOLEAN DEFAULT TRUE,
    role_id              INT DEFAULT 1,
    is_active            BOOLEAN DEFAULT TRUE,
    created_at           TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS Site_Access_ID (
    application_id          SERIAL PRIMARY KEY,
    full_name               VARCHAR(100) NOT NULL,
    job_title               VARCHAR(100),
    employee_number         VARCHAR(50),
    organization            VARCHAR(100) NOT NULL,
    subcontractor           VARCHAR(100),
    gov_id_type             VARCHAR(20),
    gov_id_number           VARCHAR(50),
    blood_type              VARCHAR(5),
    mobile_number           VARCHAR(20),
    whatsapp_number         VARCHAR(20),
    is_third_party_sponsor  BOOLEAN DEFAULT FALSE,
    workflow_state          VARCHAR(50) DEFAULT 'Pending Security Clearance',
    created_at              TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    submitted_by_user_id    INT REFERENCES App_Users(user_id)
);
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS email VARCHAR(150);
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS workflow_type VARCHAR(30) DEFAULT 'security_only';
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS form_variant VARCHAR(20) DEFAULT 'full';
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS date_joined DATE;
ALTER TABLE Site_Access_ID ADD COLUMN IF NOT EXISTS reports_to VARCHAR(100);
ALTER TABLE Site_Access_ID ALTER COLUMN workflow_state SET DEFAULT 'Pending Security Clearance';
ALTER TABLE Site_Access_ID ALTER COLUMN job_title       DROP NOT NULL;
ALTER TABLE Site_Access_ID ALTER COLUMN employee_number DROP NOT NULL;
ALTER TABLE Site_Access_ID ALTER COLUMN gov_id_type     DROP NOT NULL;
ALTER TABLE Site_Access_ID ALTER COLUMN gov_id_number   DROP NOT NULL;
ALTER TABLE Site_Access_ID ALTER COLUMN blood_type      DROP NOT NULL;
ALTER TABLE Site_Access_ID ALTER COLUMN mobile_number   DROP NOT NULL;
ALTER TABLE Site_Access_ID ALTER COLUMN whatsapp_number DROP NOT NULL;

CREATE TABLE IF NOT EXISTS Workflow_History (
    history_id            SERIAL PRIMARY KEY,
    application_id        INT NOT NULL,
    action_taken          VARCHAR(50) NOT NULL,
    previous_state        VARCHAR(50),
    new_state             VARCHAR(50) NOT NULL,
    acted_by_user_id      INT REFERENCES App_Users(user_id),
    action_timestamp      TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    comments              TEXT
);

CREATE TABLE IF NOT EXISTS Roles (
    role_id      SERIAL PRIMARY KEY,
    role_name    VARCHAR(50) UNIQUE NOT NULL,
    description  VARCHAR(255)
);

CREATE TABLE IF NOT EXISTS Role_Permissions (
    role_id          INT NOT NULL REFERENCES Roles(role_id) ON DELETE CASCADE,
    permission_key   VARCHAR(50) NOT NULL,
    PRIMARY KEY (role_id, permission_key)
);

CREATE TABLE IF NOT EXISTS Attachments (
    attachment_id        SERIAL PRIMARY KEY,
    application_id       INT NOT NULL REFERENCES Site_Access_ID(application_id) ON DELETE CASCADE,
    attachment_type      VARCHAR(50) NOT NULL,
    file_path            VARCHAR(500) NOT NULL,
    original_filename    VARCHAR(255) NOT NULL,
    mime_type            VARCHAR(100),
    file_size_bytes      BIGINT,
    uploaded_at          TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    uploaded_by_user_id  INT REFERENCES App_Users(user_id)
);

CREATE TABLE IF NOT EXISTS Notifications (
    notification_id    SERIAL PRIMARY KEY,
    application_id     INT REFERENCES Site_Access_ID(application_id) ON DELETE CASCADE,
    channel            VARCHAR(20) NOT NULL,
    recipient          VARCHAR(150) NOT NULL,
    subject            VARCHAR(255),
    body               TEXT NOT NULL,
    delivery_status    VARCHAR(30) NOT NULL,
    error_message      TEXT,
    created_at         TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    sent_at            TIMESTAMP WITH TIME ZONE
);

CREATE TABLE IF NOT EXISTS Groups (
    group_id      SERIAL PRIMARY KEY,
    group_name    VARCHAR(80) UNIQUE NOT NULL,
    description   VARCHAR(255),
    created_at    TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS User_Group_Members (
    user_id    INT NOT NULL REFERENCES App_Users(user_id)  ON DELETE CASCADE,
    group_id   INT NOT NULL REFERENCES Groups(group_id)    ON DELETE CASCADE,
    added_at   TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    PRIMARY KEY (user_id, group_id)
);

CREATE TABLE IF NOT EXISTS Submitter_Organizations (
    organization_id   SERIAL PRIMARY KEY,
    organization_name VARCHAR(100) UNIQUE NOT NULL,
    submitter_group   VARCHAR(20) NOT NULL,
    description       VARCHAR(255),
    is_active         BOOLEAN DEFAULT TRUE,
    created_at        TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS External_Submitters (
    submitter_id        SERIAL PRIMARY KEY,
    full_name           VARCHAR(100) NOT NULL,
    email               VARCHAR(150) UNIQUE NOT NULL,
    organization_id     INT NOT NULL REFERENCES Submitter_Organizations(organization_id),
    hashed_password     VARCHAR(255) NOT NULL,
    must_change_password BOOLEAN DEFAULT TRUE,
    is_active           BOOLEAN DEFAULT TRUE,
    last_login_at       TIMESTAMP WITH TIME ZONE,
    created_at          TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    created_by_user_id  INT REFERENCES App_Users(user_id)
);

CREATE TABLE IF NOT EXISTS System_Settings (
    setting_key   VARCHAR(50) PRIMARY KEY,
    setting_value VARCHAR(255) NOT NULL,
    description   VARCHAR(255),
    updated_at    TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    updated_by_user_id INT REFERENCES App_Users(user_id)
);
"""

PERMISSION_CATALOG = [
    ("submit_application",     "Can submit a new ID application"),
    ("approve_security",       "Can approve/reject at Security stage"),
    ("approve_hr",             "Can approve/reject at HR stage (legacy / standard workflow)"),
    ("view_all_applications",  "Can see every application"),
    ("edit_application",       "Can edit application fields while still pending"),
    ("manage_users",           "Can administer staff users, roles, and groups"),
    ("manage_external_users",  "Can create/manage external submitter accounts"),
    ("manage_system_settings", "Can change system settings (workflow mode, etc.)"),
]


SEED_SQL = """
INSERT INTO Roles (role_id, role_name, description) VALUES
    (1, 'Submitter',          'Internal staff who can submit applications'),
    (2, 'HR Verification',    'HR approver (used in standard workflow)'),
    (3, 'Security Clearance', 'Approves applications at the Security stage'),
    (4, 'Reviewer',           'General reviewer (read-only)'),
    (5, 'Administrator',      'Full system administrator')
ON CONFLICT (role_id) DO NOTHING;

SELECT setval(pg_get_serial_sequence('roles','role_id'),
              GREATEST((SELECT MAX(role_id) FROM Roles), 1));

INSERT INTO Role_Permissions (role_id, permission_key) VALUES
    (1, 'submit_application'),
    (2, 'approve_hr'),
    (2, 'view_all_applications'),
    (3, 'approve_security'),
    (3, 'view_all_applications'),
    (3, 'edit_application'),
    (4, 'view_all_applications'),
    (5, 'submit_application'),
    (5, 'approve_security'),
    (5, 'approve_hr'),
    (5, 'view_all_applications'),
    (5, 'edit_application'),
    (5, 'manage_users'),
    (5, 'manage_external_users'),
    (5, 'manage_system_settings')
ON CONFLICT (role_id, permission_key) DO NOTHING;

INSERT INTO App_Users (full_name, username, hashed_password, role_id, force_password_change) VALUES
    ('Omar Abdulaziz Almaqhawi','omar.almaqhawi',        'temp_hash_123', 5, TRUE),
    ('Abdulkader Abdullatif',   'abdulkader.abdullatif', 'temp_hash_123', 3, TRUE),
    ('Hind Mana Alahmari',      'hind.alahmari',         'temp_hash_123', 4, TRUE),
    ('Badriah Rizq Allah',      'badriah.rizqallah',     'temp_hash_123', 1, TRUE)
ON CONFLICT (username) DO NOTHING;

INSERT INTO Submitter_Organizations (organization_name, submitter_group, description) VALUES
    ('SEVEN',                'group_1', 'Client'),
    ('ELLISDON',             'group_1', 'Project Management Consultant (PMC)'),
    ('SBG',                  'group_1', 'Main Contractor'),
    ('AECOM',                'group_1', 'Lead Design Consultant (LDC)'),
    ('ATKINS',               'group_1', 'Lead Design Consultant (LDC)'),
    ('Subcontractor (other)','group_2', 'Catch-all for any subcontractor not separately listed')
ON CONFLICT (organization_name) DO NOTHING;

INSERT INTO System_Settings (setting_key, setting_value, description) VALUES
    ('workflow_mode', 'security_only',
     'security_only = single Security approval; standard = HR then Security')
ON CONFLICT (setting_key) DO NOTHING;
"""

# v0.7 migration: use full_name lookup instead of brittle user_id hardcode.
# Idempotent: re-running does nothing once Omar is already Administrator.
MIGRATION_SQL = """
UPDATE App_Users
   SET role_id = 5
 WHERE full_name = 'Omar Abdulaziz Almaqhawi'
   AND role_id <> 5;

WITH affected AS (
    SELECT application_id FROM Site_Access_ID
     WHERE workflow_state = 'Pending HR Verification'
)
INSERT INTO Workflow_History
    (application_id, action_taken, previous_state, new_state, acted_by_user_id, comments)
SELECT application_id, 'Workflow Migration', 'Pending HR Verification',
       'Pending Security Clearance', NULL,
       'Auto-migrated to security-only workflow'
  FROM affected;

UPDATE Site_Access_ID
   SET workflow_state = 'Pending Security Clearance'
 WHERE workflow_state = 'Pending HR Verification';
"""


@app.post("/api/v1/admin/init-db")
async def init_db():
    conn = await get_conn()
    try:
        await conn.execute(SCHEMA_SQL)
        await conn.execute(SEED_SQL)
        await conn.execute(MIGRATION_SQL)
        for key, _desc in PERMISSION_CATALOG:
            await conn.execute(
                """INSERT INTO Role_Permissions (role_id, permission_key)
                   VALUES (5, $1) ON CONFLICT DO NOTHING""",
                key,
            )
    finally:
        await conn.close()
    return {"status": "ok", "message": "Schema, seeds, and migrations applied."}


# ============================================================
#  Models
# ============================================================
class StatusUpdate(BaseModel):
    new_state: str
    acted_by_user_id: int
    comments: Optional[str] = None


class UserCreate(BaseModel):
    full_name: str
    username: str
    role_id: int
    is_active: bool = True


class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    role_id: Optional[int] = None
    is_active: Optional[bool] = None


class RoleCreate(BaseModel):
    role_name: str
    description: Optional[str] = None
    permissions: List[str] = []


class RolePermissionsUpdate(BaseModel):
    permissions: List[str]


class GroupCreate(BaseModel):
    group_name: str
    description: Optional[str] = None
    member_user_ids: List[int] = []


class GroupMembersUpdate(BaseModel):
    member_user_ids: List[int]


class ExternalSubmitterCreate(BaseModel):
    full_name: str
    email: str
    organization_id: int


class ApplicationFieldsUpdate(BaseModel):
    full_name:        Optional[str] = None
    job_title:        Optional[str] = None
    employee_number:  Optional[str] = None
    organization:     Optional[str] = None
    subcontractor:    Optional[str] = None
    gov_id_type:      Optional[str] = None
    gov_id_number:    Optional[str] = None
    blood_type:       Optional[str] = None
    mobile_number:    Optional[str] = None
    whatsapp_number:  Optional[str] = None
    email:            Optional[str] = None
    is_third_party_sponsor: Optional[bool] = None
    date_joined:      Optional[str] = None
    reports_to:       Optional[str] = None
    edited_by_user_id: int


class SettingUpdate(BaseModel):
    setting_value: str


# ============================================================
#  Permission helpers
# ============================================================
async def fetch_user_permissions(conn, user_id):
    rows = await conn.fetch(
        """SELECT rp.permission_key
             FROM Role_Permissions rp
             JOIN App_Users u ON u.role_id = rp.role_id
            WHERE u.user_id = $1""",
        user_id,
    )
    return [r["permission_key"] for r in rows]


async def get_workflow_mode(conn) -> str:
    row = await conn.fetchrow(
        "SELECT setting_value FROM System_Settings WHERE setting_key = 'workflow_mode'"
    )
    return row["setting_value"] if row else "security_only"


async def required_permission_for_state_change(conn, current_state: str, new_state: str) -> Optional[str]:
    mode = await get_workflow_mode(conn)
    if mode == "standard":
        if current_state == "Pending HR Verification" and new_state in ("Pending Security Clearance", "Rejected"):
            return "approve_hr"
        if current_state == "Pending Security Clearance" and new_state in ("Approved", "Rejected"):
            return "approve_security"
    else:
        if current_state == "Pending Security Clearance" and new_state in ("Approved", "Rejected"):
            return "approve_security"
        if current_state == "Pending HR Verification" and new_state in ("Pending Security Clearance", "Rejected"):
            return "approve_hr"
    return None


async def require_permission(conn, acting_user_id, permission):
    if not acting_user_id:
        raise HTTPException(401, "Missing X-Acting-User-Id header")
    perms = await fetch_user_permissions(conn, acting_user_id)
    if permission not in perms:
        raise HTTPException(403, f"User lacks permission '{permission}'")


def acting_user_id_from_header(x_acting_user_id):
    if not x_acting_user_id:
        return None
    try:
        return int(x_acting_user_id)
    except ValueError:
        return None


def parse_date_or_none(s: Optional[str]) -> Optional[date]:
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, f"Invalid date format '{s}' (expected YYYY-MM-DD)")


# ============================================================
#  Password helpers
# ============================================================
def hash_password(plaintext: str) -> str:
    return bcrypt.hashpw(plaintext.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def generate_password() -> str:
    return secrets.token_urlsafe(9)


# ============================================================
#  Notifications
# ============================================================
def _build_rejection_email_body(applicant_name, app_id, reason):
    return (
        f"Dear {applicant_name},\n\n"
        f"Your Site Access ID application (#{app_id}) has been reviewed and "
        f"unfortunately could not be approved at this time.\n\n"
        f"Reason provided by the reviewer:\n{reason}\n\n"
        f"If you believe this was made in error or you would like to re-apply with "
        f"corrected details, please contact your site administrator.\n\n"
        f"Regards,\nTa'awon P3 Security Management System"
    )


def _build_rejection_whatsapp_body(applicant_name, app_id, reason):
    return (
        f"Hello {applicant_name}, your Site Access ID application "
        f"(#{app_id}) was not approved.\nReason: {reason}\n"
        f"Please contact your site administrator for next steps."
    )


def _build_credentials_email_body(full_name, email, password, organization_name):
    return (
        f"Dear {full_name},\n\n"
        f"An account has been created for you on the Ta'awon P3 Security Management System "
        f"so that you can submit Site Access ID applications on behalf of {organization_name}.\n\n"
        f"Your login credentials are:\n"
        f"  Username: {email}\n"
        f"  Password: {password}\n\n"
        f"You will be asked to change your password on first login.\n\n"
        f"NOTE: The login portal is currently being prepared. You will receive a separate "
        f"email with the portal URL once it is available.\n\n"
        f"Regards,\nTa'awon P3 Security Management System"
    )


async def _log_notification(conn, app_id, channel, recipient, subject, body,
                            status, error_message=None, sent_at=None):
    await conn.execute(
        """INSERT INTO Notifications
              (application_id, channel, recipient, subject, body,
               delivery_status, error_message, sent_at)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8)""",
        app_id, channel, recipient, subject, body,
        status, error_message, sent_at,
    )


def _send_email_smtp(to_addr, subject, body):
    if not SMTP_ENABLED:
        return False, "SMTP not configured (SMTP_USERNAME/SMTP_PASSWORD missing in .env)"
    try:
        msg = MIMEMultipart()
        msg["From"] = SMTP_FROM
        msg["To"] = to_addr
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain", "utf-8"))
        context = ssl.create_default_context()
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as server:
            server.starttls(context=context)
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(SMTP_FROM, [to_addr], msg.as_string())
        return True, None
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


async def _handle_rejection_notifications(conn, app_record, reason):
    app_id = app_record["application_id"]
    name = app_record["full_name"]
    email = app_record["email"]
    whatsapp = app_record["whatsapp_number"] or app_record["mobile_number"]

    if not email:
        await _log_notification(conn, app_id, "email", "(unknown)", None,
                                "(skipped - no email on file)", "skipped_no_recipient")
    else:
        subject = f"Site Access ID Application #{app_id} - Update"
        body = _build_rejection_email_body(name, app_id, reason)
        success, err = _send_email_smtp(email, subject, body)
        await _log_notification(conn, app_id, "email", email, subject, body,
                                "sent" if success else "failed",
                                err, datetime.utcnow() if success else None)

    if not whatsapp:
        await _log_notification(conn, app_id, "whatsapp", "(unknown)", None,
                                "(skipped - no WhatsApp number)", "skipped_no_recipient")
    else:
        body = _build_rejection_whatsapp_body(name, app_id, reason)
        await _log_notification(conn, app_id, "whatsapp", whatsapp, None, body,
                                "pending_whatsapp_api")


# ============================================================
#  Endpoints
# ============================================================
@app.get("/")
async def root():
    return {
        "service": "Ta'awon P3 SMS API",
        "version": app.version,
        "smtp_enabled": SMTP_ENABLED,
        "cors_origins": CORS_ORIGINS,
    }


@app.get("/api/v1/users")
async def list_users():
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT u.user_id, u.full_name, u.username, u.role_id,
                      r.role_name, u.is_active
                 FROM App_Users u
            LEFT JOIN Roles r ON r.role_id = u.role_id
                WHERE u.is_active = TRUE
             ORDER BY u.user_id"""
        )
        return {"users": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.get("/api/v1/users/{user_id}/permissions")
async def get_user_permissions(user_id: int):
    conn = await get_conn()
    try:
        user = await conn.fetchrow(
            """SELECT u.user_id, u.full_name, u.role_id, r.role_name
                 FROM App_Users u
            LEFT JOIN Roles r ON r.role_id = u.role_id
                WHERE u.user_id = $1""",
            user_id,
        )
        if not user:
            raise HTTPException(404, "User not found")
        perms = await fetch_user_permissions(conn, user_id)
        return {**dict(user), "permissions": perms}
    finally:
        await conn.close()


@app.get("/api/v1/permissions")
async def list_permission_catalog():
    return {"permissions": [{"key": k, "description": d} for k, d in PERMISSION_CATALOG]}


# ---------- Admin: Users ----------
@app.get("/api/v1/admin/users")
async def admin_list_users(x_acting_user_id: Optional[str] = Header(None)):
    conn = await get_conn()
    try:
        await require_permission(conn, acting_user_id_from_header(x_acting_user_id), "manage_users")
        rows = await conn.fetch(
            """SELECT u.user_id, u.full_name, u.username, u.role_id, r.role_name, u.is_active, u.created_at
                 FROM App_Users u
            LEFT JOIN Roles r ON r.role_id = u.role_id
             ORDER BY u.user_id"""
        )
        return {"users": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.post("/api/v1/admin/users")
async def admin_create_user(payload: UserCreate, x_acting_user_id: Optional[str] = Header(None)):
    conn = await get_conn()
    try:
        await require_permission(conn, acting_user_id_from_header(x_acting_user_id), "manage_users")
        try:
            row = await conn.fetchrow(
                """INSERT INTO App_Users (full_name, username, hashed_password, role_id, is_active)
                   VALUES ($1, $2, 'temp_hash_123', $3, $4)
                RETURNING user_id, full_name, username, role_id, is_active""",
                payload.full_name.strip(), payload.username.strip(),
                payload.role_id, payload.is_active,
            )
        except asyncpg.UniqueViolationError:
            raise HTTPException(409, f"Username '{payload.username}' already exists")
        return {"status": "ok", "user": dict(row)}
    finally:
        await conn.close()


@app.put("/api/v1/admin/users/{user_id}")
async def admin_update_user(user_id: int, payload: UserUpdate,
                            x_acting_user_id: Optional[str] = Header(None)):
    conn = await get_conn()
    try:
        await require_permission(conn, acting_user_id_from_header(x_acting_user_id), "manage_users")
        sets, vals = [], []
        if payload.full_name is not None:
            sets.append(f"full_name = ${len(vals) + 1}"); vals.append(payload.full_name.strip())
        if payload.role_id is not None:
            sets.append(f"role_id = ${len(vals) + 1}"); vals.append(payload.role_id)
        if payload.is_active is not None:
            sets.append(f"is_active = ${len(vals) + 1}"); vals.append(payload.is_active)
        if not sets:
            raise HTTPException(400, "No fields to update")
        vals.append(user_id)
        q = f"UPDATE App_Users SET {', '.join(sets)} WHERE user_id = ${len(vals)} RETURNING user_id"
        row = await conn.fetchrow(q, *vals)
        if not row:
            raise HTTPException(404, "User not found")
        return {"status": "ok", "user_id": row["user_id"]}
    finally:
        await conn.close()


@app.delete("/api/v1/admin/users/{user_id}")
async def admin_deactivate_user(user_id: int, x_acting_user_id: Optional[str] = Header(None)):
    conn = await get_conn()
    try:
        acting = acting_user_id_from_header(x_acting_user_id)
        await require_permission(conn, acting, "manage_users")
        if acting == user_id:
            raise HTTPException(400, "You cannot deactivate yourself")
        row = await conn.fetchrow(
            "UPDATE App_Users SET is_active = FALSE WHERE user_id = $1 RETURNING user_id",
            user_id,
        )
        if not row:
            raise HTTPException(404, "User not found")
        return {"status": "ok"}
    finally:
        await conn.close()


# ---------- Admin: Roles & Permissions ----------
@app.get("/api/v1/admin/roles")
async def admin_list_roles(x_acting_user_id: Optional[str] = Header(None)):
    conn = await get_conn()
    try:
        await require_permission(conn, acting_user_id_from_header(x_acting_user_id), "manage_users")
        roles = await conn.fetch("SELECT role_id, role_name, description FROM Roles ORDER BY role_id")
        out = []
        for r in roles:
            perms = await conn.fetch(
                "SELECT permission_key FROM Role_Permissions WHERE role_id = $1 ORDER BY permission_key",
                r["role_id"],
            )
            out.append({**dict(r), "permissions": [p["permission_key"] for p in perms]})
        return {"roles": out}
    finally:
        await conn.close()


@app.post("/api/v1/admin/roles")
async def admin_create_role(payload: RoleCreate, x_acting_user_id: Optional[str] = Header(None)):
    valid_keys = {k for k, _ in PERMISSION_CATALOG}
    bad = [p for p in payload.permissions if p not in valid_keys]
    if bad:
        raise HTTPException(400, f"Unknown permission keys: {bad}")
    conn = await get_conn()
    try:
        await require_permission(conn, acting_user_id_from_header(x_acting_user_id), "manage_users")
        async with conn.transaction():
            try:
                row = await conn.fetchrow(
                    """INSERT INTO Roles (role_name, description) VALUES ($1, $2)
                    RETURNING role_id, role_name, description""",
                    payload.role_name.strip(), payload.description,
                )
            except asyncpg.UniqueViolationError:
                raise HTTPException(409, f"Role '{payload.role_name}' already exists")
            for key in payload.permissions:
                await conn.execute(
                    "INSERT INTO Role_Permissions (role_id, permission_key) VALUES ($1, $2)",
                    row["role_id"], key,
                )
        return {"status": "ok", "role": dict(row), "permissions": payload.permissions}
    finally:
        await conn.close()


@app.put("/api/v1/admin/roles/{role_id}/permissions")
async def admin_update_role_permissions(role_id: int, payload: RolePermissionsUpdate,
                                        x_acting_user_id: Optional[str] = Header(None)):
    valid_keys = {k for k, _ in PERMISSION_CATALOG}
    bad = [p for p in payload.permissions if p not in valid_keys]
    if bad:
        raise HTTPException(400, f"Unknown permission keys: {bad}")
    conn = await get_conn()
    try:
        await require_permission(conn, acting_user_id_from_header(x_acting_user_id), "manage_users")
        role = await conn.fetchrow("SELECT role_id FROM Roles WHERE role_id = $1", role_id)
        if not role:
            raise HTTPException(404, "Role not found")
        async with conn.transaction():
            await conn.execute("DELETE FROM Role_Permissions WHERE role_id = $1", role_id)
            for key in payload.permissions:
                await conn.execute(
                    "INSERT INTO Role_Permissions (role_id, permission_key) VALUES ($1, $2)",
                    role_id, key,
                )
        return {"status": "ok"}
    finally:
        await conn.close()


# ---------- Admin: Groups ----------
@app.get("/api/v1/admin/groups")
async def admin_list_groups(x_acting_user_id: Optional[str] = Header(None)):
    conn = await get_conn()
    try:
        await require_permission(conn, acting_user_id_from_header(x_acting_user_id), "manage_users")
        groups = await conn.fetch("SELECT group_id, group_name, description, created_at FROM Groups ORDER BY group_id")
        out = []
        for g in groups:
            members = await conn.fetch(
                """SELECT u.user_id, u.full_name, u.username
                     FROM User_Group_Members m
                     JOIN App_Users u ON u.user_id = m.user_id
                    WHERE m.group_id = $1
                 ORDER BY u.full_name""",
                g["group_id"],
            )
            out.append({**dict(g), "members": [dict(m) for m in members]})
        return {"groups": out}
    finally:
        await conn.close()


@app.post("/api/v1/admin/groups")
async def admin_create_group(payload: GroupCreate, x_acting_user_id: Optional[str] = Header(None)):
    conn = await get_conn()
    try:
        await require_permission(conn, acting_user_id_from_header(x_acting_user_id), "manage_users")
        async with conn.transaction():
            try:
                row = await conn.fetchrow(
                    """INSERT INTO Groups (group_name, description) VALUES ($1, $2)
                    RETURNING group_id, group_name, description""",
                    payload.group_name.strip(), payload.description,
                )
            except asyncpg.UniqueViolationError:
                raise HTTPException(409, f"Group '{payload.group_name}' already exists")
            for uid in payload.member_user_ids:
                await conn.execute(
                    "INSERT INTO User_Group_Members (user_id, group_id) VALUES ($1, $2)",
                    uid, row["group_id"],
                )
        return {"status": "ok", "group": dict(row)}
    finally:
        await conn.close()


@app.put("/api/v1/admin/groups/{group_id}/members")
async def admin_update_group_members(group_id: int, payload: GroupMembersUpdate,
                                     x_acting_user_id: Optional[str] = Header(None)):
    conn = await get_conn()
    try:
        await require_permission(conn, acting_user_id_from_header(x_acting_user_id), "manage_users")
        g = await conn.fetchrow("SELECT group_id FROM Groups WHERE group_id = $1", group_id)
        if not g:
            raise HTTPException(404, "Group not found")
        async with conn.transaction():
            await conn.execute("DELETE FROM User_Group_Members WHERE group_id = $1", group_id)
            for uid in payload.member_user_ids:
                await conn.execute(
                    "INSERT INTO User_Group_Members (user_id, group_id) VALUES ($1, $2)",
                    uid, group_id,
                )
        return {"status": "ok"}
    finally:
        await conn.close()


@app.delete("/api/v1/admin/groups/{group_id}")
async def admin_delete_group(group_id: int, x_acting_user_id: Optional[str] = Header(None)):
    conn = await get_conn()
    try:
        await require_permission(conn, acting_user_id_from_header(x_acting_user_id), "manage_users")
        row = await conn.fetchrow("DELETE FROM Groups WHERE group_id = $1 RETURNING group_id", group_id)
        if not row:
            raise HTTPException(404, "Group not found")
        return {"status": "ok"}
    finally:
        await conn.close()


# ---------- Submitter Organizations ----------
@app.get("/api/v1/organizations")
async def list_organizations():
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT organization_id, organization_name, submitter_group, description, is_active
                 FROM Submitter_Organizations
                WHERE is_active = TRUE
             ORDER BY submitter_group, organization_name"""
        )
        return {"organizations": [dict(r) for r in rows]}
    finally:
        await conn.close()


# ---------- Admin: External Submitters ----------
@app.get("/api/v1/admin/external-submitters")
async def admin_list_external_submitters(x_acting_user_id: Optional[str] = Header(None)):
    conn = await get_conn()
    try:
        await require_permission(conn, acting_user_id_from_header(x_acting_user_id), "manage_external_users")
        rows = await conn.fetch(
            """SELECT es.submitter_id, es.full_name, es.email, es.is_active,
                      es.must_change_password, es.last_login_at, es.created_at,
                      so.organization_id, so.organization_name, so.submitter_group
                 FROM External_Submitters es
            LEFT JOIN Submitter_Organizations so ON so.organization_id = es.organization_id
             ORDER BY es.submitter_id DESC"""
        )
        return {"external_submitters": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.post("/api/v1/admin/external-submitters")
async def admin_create_external_submitter(
    payload: ExternalSubmitterCreate, x_acting_user_id: Optional[str] = Header(None)
):
    conn = await get_conn()
    try:
        acting = acting_user_id_from_header(x_acting_user_id)
        await require_permission(conn, acting, "manage_external_users")
        org = await conn.fetchrow(
            "SELECT organization_id, organization_name FROM Submitter_Organizations WHERE organization_id = $1",
            payload.organization_id,
        )
        if not org:
            raise HTTPException(400, "Unknown organization_id")
        plain_password = generate_password()
        hashed = hash_password(plain_password)
        try:
            row = await conn.fetchrow(
                """INSERT INTO External_Submitters
                      (full_name, email, organization_id, hashed_password, created_by_user_id)
                   VALUES ($1, $2, $3, $4, $5)
                RETURNING submitter_id, full_name, email, organization_id""",
                payload.full_name.strip(), payload.email.strip(),
                payload.organization_id, hashed, acting,
            )
        except asyncpg.UniqueViolationError:
            raise HTTPException(409, f"An external submitter with email '{payload.email}' already exists")
        subject = "Your Ta'awon P3 SMS submitter account"
        body = _build_credentials_email_body(
            payload.full_name.strip(), payload.email.strip(),
            plain_password, org["organization_name"],
        )
        success, err = _send_email_smtp(payload.email.strip(), subject, body)
        await _log_notification(
            conn, None, "email", payload.email.strip(), subject, body,
            "sent" if success else "failed",
            err, datetime.utcnow() if success else None,
        )
        return {
            "status": "ok",
            "external_submitter": dict(row),
            "credential_email_sent": success,
            "credential_email_error": err,
            "temporary_password_for_manual_share": plain_password if not success else None,
        }
    finally:
        await conn.close()


@app.put("/api/v1/admin/external-submitters/{submitter_id}/regenerate-password")
async def admin_regenerate_external_password(
    submitter_id: int, x_acting_user_id: Optional[str] = Header(None)
):
    conn = await get_conn()
    try:
        await require_permission(conn, acting_user_id_from_header(x_acting_user_id), "manage_external_users")
        sub = await conn.fetchrow(
            """SELECT es.full_name, es.email, so.organization_name
                 FROM External_Submitters es
            LEFT JOIN Submitter_Organizations so ON so.organization_id = es.organization_id
                WHERE es.submitter_id = $1""",
            submitter_id,
        )
        if not sub:
            raise HTTPException(404, "External submitter not found")
        plain_password = generate_password()
        hashed = hash_password(plain_password)
        await conn.execute(
            """UPDATE External_Submitters
                  SET hashed_password = $1, must_change_password = TRUE
                WHERE submitter_id = $2""",
            hashed, submitter_id,
        )
        subject = "Ta'awon P3 SMS - your new password"
        body = _build_credentials_email_body(
            sub["full_name"], sub["email"], plain_password,
            sub["organization_name"] or "(unspecified organization)",
        )
        success, err = _send_email_smtp(sub["email"], subject, body)
        await _log_notification(
            conn, None, "email", sub["email"], subject, body,
            "sent" if success else "failed",
            err, datetime.utcnow() if success else None,
        )
        return {
            "status": "ok",
            "credential_email_sent": success,
            "credential_email_error": err,
            "temporary_password_for_manual_share": plain_password if not success else None,
        }
    finally:
        await conn.close()


@app.delete("/api/v1/admin/external-submitters/{submitter_id}")
async def admin_deactivate_external_submitter(
    submitter_id: int, x_acting_user_id: Optional[str] = Header(None)
):
    conn = await get_conn()
    try:
        await require_permission(conn, acting_user_id_from_header(x_acting_user_id), "manage_external_users")
        row = await conn.fetchrow(
            "UPDATE External_Submitters SET is_active = FALSE WHERE submitter_id = $1 RETURNING submitter_id",
            submitter_id,
        )
        if not row:
            raise HTTPException(404, "External submitter not found")
        return {"status": "ok"}
    finally:
        await conn.close()


# ---------- Admin: System Settings ----------
@app.get("/api/v1/settings")
async def public_settings():
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            "SELECT setting_key, setting_value FROM System_Settings WHERE setting_key IN ('workflow_mode')"
        )
        return {"settings": {r["setting_key"]: r["setting_value"] for r in rows}}
    finally:
        await conn.close()


@app.get("/api/v1/admin/settings")
async def admin_list_settings(x_acting_user_id: Optional[str] = Header(None)):
    conn = await get_conn()
    try:
        await require_permission(conn, acting_user_id_from_header(x_acting_user_id), "manage_system_settings")
        rows = await conn.fetch(
            "SELECT setting_key, setting_value, description, updated_at FROM System_Settings ORDER BY setting_key"
        )
        return {"settings": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.put("/api/v1/admin/settings/{setting_key}")
async def admin_update_setting(setting_key: str, payload: SettingUpdate,
                               x_acting_user_id: Optional[str] = Header(None)):
    conn = await get_conn()
    try:
        acting = acting_user_id_from_header(x_acting_user_id)
        await require_permission(conn, acting, "manage_system_settings")
        if setting_key == "workflow_mode" and payload.setting_value not in ("security_only", "standard"):
            raise HTTPException(400, "workflow_mode must be 'security_only' or 'standard'")
        row = await conn.fetchrow(
            """UPDATE System_Settings
                  SET setting_value = $1, updated_at = NOW(), updated_by_user_id = $2
                WHERE setting_key = $3
            RETURNING setting_key, setting_value""",
            payload.setting_value, acting, setting_key,
        )
        if not row:
            raise HTTPException(404, f"Unknown setting '{setting_key}'")
        return {"status": "ok", **dict(row)}
    finally:
        await conn.close()


# ============================================================
#  Endpoints - applications
# ============================================================
def _validate_upload(upload):
    if upload.content_type:
        ok = (upload.content_type in ALLOWED_MIME_EXACT or
              any(upload.content_type.startswith(p) for p in ALLOWED_MIME_PREFIXES))
        if not ok:
            raise HTTPException(415, f"Unsupported file type: {upload.content_type}")


async def _save_upload(upload, app_id, attachment_type):
    _validate_upload(upload)
    contents = await upload.read()
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"File '{upload.filename}' exceeds {MAX_UPLOAD_BYTES // (1024*1024)}MB limit")
    app_dir = UPLOADS_DIR / f"app_{app_id}"
    app_dir.mkdir(exist_ok=True)
    original_name = Path(upload.filename or "file").name
    ext = Path(original_name).suffix.lower()
    stored_name = f"{attachment_type}_{uuid.uuid4().hex[:12]}{ext}"
    dest_path = app_dir / stored_name
    with open(dest_path, "wb") as f:
        f.write(contents)
    rel_path = dest_path.relative_to(UPLOADS_DIR).as_posix()
    return {
        "file_path": rel_path,
        "original_filename": original_name,
        "mime_type": upload.content_type,
        "file_size_bytes": len(contents),
    }


@app.post("/api/v1/applications/submit")
async def submit_application(
    full_name: str  = Form(...),
    organization: str  = Form(...),
    email: Optional[str] = Form(None),
    submitted_by_user_id: int = Form(...),
    form_variant: str = Form("full"),
    job_title: Optional[str] = Form(None),
    employee_number: Optional[str] = Form(None),
    subcontractor: Optional[str] = Form(None),
    gov_id_type: Optional[str] = Form(None),
    gov_id_number: Optional[str] = Form(None),
    blood_type: Optional[str] = Form(None),
    mobile_number: Optional[str] = Form(None),
    whatsapp_number: Optional[str] = Form(None),
    is_third_party_sponsor: bool = Form(False),
    date_joined: Optional[str] = Form(None),
    reports_to: Optional[str] = Form(None),
    id_copy: Optional[UploadFile] = File(None),
    insurance: Optional[UploadFile] = File(None),
    photo: Optional[UploadFile] = File(None),
    legal_agreement: Optional[UploadFile] = File(None),
):
    conn = await get_conn()
    try:
        mode = await get_workflow_mode(conn)
        starting_state = "Pending HR Verification" if mode == "standard" else "Pending Security Clearance"
        date_joined_val = parse_date_or_none(date_joined)

        async with conn.transaction():
            row = await conn.fetchrow(
                """INSERT INTO Site_Access_ID
                      (full_name, job_title, employee_number, organization,
                       subcontractor, gov_id_type, gov_id_number, blood_type,
                       mobile_number, whatsapp_number, email,
                       is_third_party_sponsor, submitted_by_user_id,
                       workflow_state, workflow_type,
                       form_variant, date_joined, reports_to)
                   VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18)
                RETURNING application_id, workflow_state""",
                full_name, job_title, employee_number, organization,
                subcontractor, gov_id_type, gov_id_number, blood_type,
                mobile_number, whatsapp_number, email,
                is_third_party_sponsor, submitted_by_user_id,
                starting_state, mode,
                form_variant, date_joined_val, reports_to,
            )
            app_id = row["application_id"]

            uploads = {"id_copy": id_copy, "insurance": insurance, "photo": photo, "legal_agreement": legal_agreement}
            for kind, upload in uploads.items():
                if upload is None or upload.filename in (None, ""):
                    continue
                meta = await _save_upload(upload, app_id, kind)
                await conn.execute(
                    """INSERT INTO Attachments
                          (application_id, attachment_type, file_path, original_filename,
                           mime_type, file_size_bytes, uploaded_by_user_id)
                       VALUES ($1,$2,$3,$4,$5,$6,$7)""",
                    app_id, kind, meta["file_path"], meta["original_filename"],
                    meta["mime_type"], meta["file_size_bytes"], submitted_by_user_id,
                )

            await conn.execute(
                """INSERT INTO Workflow_History
                      (application_id, action_taken, previous_state, new_state, acted_by_user_id, comments)
                   VALUES ($1, 'Submitted', 'Draft', $2, $3, $4)""",
                app_id, row["workflow_state"], submitted_by_user_id,
                f"Application submitted ({form_variant} form)",
            )
        return {"status": "ok", "application_id": app_id}
    finally:
        await conn.close()


@app.get("/api/v1/applications/pending")
async def list_pending():
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT application_id, full_name, organization, workflow_state, created_at
                 FROM Site_Access_ID
                WHERE workflow_state NOT IN ('Approved','Rejected','Cancelled')
             ORDER BY application_id DESC"""
        )
        return {"pending_applications": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.get("/api/v1/applications")
async def list_all_applications():
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT application_id, full_name, organization, workflow_state, created_at
                 FROM Site_Access_ID
             ORDER BY application_id DESC"""
        )
        return {"applications": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.get("/api/v1/applications/export")
async def export_applications_xlsx():
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT s.application_id, s.created_at, s.workflow_state, s.workflow_type, s.form_variant,
                      s.full_name, s.email, s.mobile_number, s.whatsapp_number,
                      s.gov_id_type, s.gov_id_number, s.blood_type,
                      s.job_title, s.employee_number, s.organization, s.subcontractor,
                      s.is_third_party_sponsor, s.date_joined, s.reports_to,
                      u.full_name AS submitter_name
                 FROM Site_Access_ID s
            LEFT JOIN App_Users u ON u.user_id = s.submitted_by_user_id
             ORDER BY s.application_id ASC"""
        )
    finally:
        await conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    headers = [
        "App ID", "Submitted At", "Status", "Workflow Type", "Form Variant",
        "Full Name", "Email", "Mobile", "WhatsApp",
        "Gov ID Type", "Gov ID Number", "Blood Type",
        "Job/Trade", "Employee #", "Organization", "Subcontractor",
        "3rd-Party Sponsor", "Date Joined", "Reports To", "Submitted By",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="00A651")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append([
            r["application_id"],
            r["created_at"].strftime("%Y-%m-%d %H:%M") if r["created_at"] else "",
            r["workflow_state"], r["workflow_type"], r["form_variant"] or "",
            r["full_name"], r["email"] or "", r["mobile_number"] or "", r["whatsapp_number"] or "",
            r["gov_id_type"] or "", r["gov_id_number"] or "", r["blood_type"] or "",
            r["job_title"] or "", r["employee_number"] or "",
            r["organization"] or "", r["subcontractor"] or "",
            "Yes" if r["is_third_party_sponsor"] else "No",
            r["date_joined"].strftime("%Y-%m-%d") if r["date_joined"] else "",
            r["reports_to"] or "",
            r["submitter_name"] or "",
        ])
    for col_letter in [chr(ord('A') + i) for i in range(len(headers))]:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in ws[col_letter])
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"taawon_p3_applications_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/v1/applications/{application_id}")
async def get_application_full(application_id: int):
    conn = await get_conn()
    try:
        app_row = await conn.fetchrow(
            """SELECT s.*, u.full_name AS submitter_name
                 FROM Site_Access_ID s
            LEFT JOIN App_Users u ON u.user_id = s.submitted_by_user_id
                WHERE s.application_id = $1""",
            application_id,
        )
        if not app_row:
            raise HTTPException(404, "Application not found")
        history_rows = await conn.fetch(
            """SELECT h.*, u.full_name AS acted_by
                 FROM Workflow_History h
            LEFT JOIN App_Users u ON u.user_id = h.acted_by_user_id
                WHERE h.application_id = $1
             ORDER BY h.action_timestamp ASC""",
            application_id,
        )
        attach_rows = await conn.fetch(
            """SELECT attachment_id, attachment_type, original_filename,
                      mime_type, file_size_bytes, uploaded_at
                 FROM Attachments
                WHERE application_id = $1
             ORDER BY attachment_id ASC""",
            application_id,
        )
        notif_rows = await conn.fetch(
            """SELECT notification_id, channel, recipient, subject, body,
                      delivery_status, error_message, created_at, sent_at
                 FROM Notifications
                WHERE application_id = $1
             ORDER BY notification_id DESC""",
            application_id,
        )
        return {
            "application":   dict(app_row),
            "history":       [dict(r) for r in history_rows],
            "attachments":   [dict(r) for r in attach_rows],
            "notifications": [dict(r) for r in notif_rows],
        }
    finally:
        await conn.close()


@app.put("/api/v1/applications/{application_id}")
async def edit_application_fields(application_id: int, payload: ApplicationFieldsUpdate,
                                  x_acting_user_id: Optional[str] = Header(None)):
    conn = await get_conn()
    try:
        editor_id = payload.edited_by_user_id
        await require_permission(conn, editor_id, "edit_application")
        app_row = await conn.fetchrow(
            "SELECT workflow_state FROM Site_Access_ID WHERE application_id = $1",
            application_id,
        )
        if not app_row:
            raise HTTPException(404, "Application not found")
        if app_row["workflow_state"] in ("Approved", "Rejected", "Cancelled"):
            raise HTTPException(409, f"Application is in '{app_row['workflow_state']}' state and cannot be edited")
        editable_fields = [
            "full_name", "job_title", "employee_number", "organization", "subcontractor",
            "gov_id_type", "gov_id_number", "blood_type", "mobile_number", "whatsapp_number",
            "email", "is_third_party_sponsor", "reports_to",
        ]
        sets, vals = [], []
        changed_summary = []
        for f in editable_fields:
            v = getattr(payload, f)
            if v is not None:
                sets.append(f"{f} = ${len(vals) + 1}")
                vals.append(v)
                changed_summary.append(f)
        if payload.date_joined is not None:
            sets.append(f"date_joined = ${len(vals) + 1}")
            vals.append(parse_date_or_none(payload.date_joined))
            changed_summary.append("date_joined")
        if not sets:
            raise HTTPException(400, "No editable fields provided")
        vals.append(application_id)
        async with conn.transaction():
            await conn.execute(
                f"UPDATE Site_Access_ID SET {', '.join(sets)} WHERE application_id = ${len(vals)}",
                *vals,
            )
            await conn.execute(
                """INSERT INTO Workflow_History
                      (application_id, action_taken, previous_state, new_state,
                       acted_by_user_id, comments)
                   VALUES ($1, 'Edit', $2, $2, $3, $4)""",
                application_id, app_row["workflow_state"], editor_id,
                f"Fields edited: {', '.join(changed_summary)}",
            )
        return {"status": "ok", "fields_changed": changed_summary}
    finally:
        await conn.close()


@app.get("/api/v1/applications/{application_id}/attachment/{attachment_id}")
async def download_attachment(application_id: int, attachment_id: int):
    conn = await get_conn()
    try:
        row = await conn.fetchrow(
            """SELECT file_path, original_filename, mime_type
                 FROM Attachments
                WHERE attachment_id = $1 AND application_id = $2""",
            attachment_id, application_id,
        )
    finally:
        await conn.close()
    if not row:
        raise HTTPException(404, "Attachment not found")
    full_path = (UPLOADS_DIR / row["file_path"]).resolve()
    try:
        full_path.relative_to(UPLOADS_DIR.resolve())
    except ValueError:
        raise HTTPException(400, "Invalid path")
    if not full_path.is_file():
        raise HTTPException(410, "File missing on disk")
    return FileResponse(
        full_path,
        media_type=row["mime_type"] or "application/octet-stream",
        filename=row["original_filename"],
    )


@app.put("/api/v1/applications/{application_id}/status")
async def update_status(application_id: int, payload: StatusUpdate):
    conn = await get_conn()
    try:
        async with conn.transaction():
            app_row = await conn.fetchrow(
                "SELECT * FROM Site_Access_ID WHERE application_id = $1",
                application_id,
            )
            if not app_row:
                raise HTTPException(404, "Application not found")
            current_state = app_row["workflow_state"]
            required = await required_permission_for_state_change(conn, current_state, payload.new_state)
            if required:
                perms = await fetch_user_permissions(conn, payload.acted_by_user_id)
                if required not in perms:
                    raise HTTPException(
                        403,
                        f"User lacks permission '{required}' (transition: {current_state} -> {payload.new_state})",
                    )
            await conn.execute(
                "UPDATE Site_Access_ID SET workflow_state = $1 WHERE application_id = $2",
                payload.new_state, application_id,
            )
            await conn.execute(
                """INSERT INTO Workflow_History
                      (application_id, action_taken, previous_state, new_state,
                       acted_by_user_id, comments)
                   VALUES ($1, $2, $3, $4, $5, $6)""",
                application_id,
                "Reject" if payload.new_state == "Rejected" else "Approve",
                current_state, payload.new_state,
                payload.acted_by_user_id, payload.comments,
            )
            if payload.new_state == "Rejected":
                reason = payload.comments or "(no reason provided)"
                await _handle_rejection_notifications(conn, app_row, reason)
        return {"status": "ok", "application_id": application_id, "new_state": payload.new_state}
    finally:
        await conn.close()


@app.get("/api/v1/history")
async def list_history():
    conn = await get_conn()
    try:
        rows = await conn.fetch(
            """SELECT h.*, u.full_name AS acted_by
                 FROM Workflow_History h
            LEFT JOIN App_Users u ON u.user_id = h.acted_by_user_id
             ORDER BY h.action_timestamp DESC
                LIMIT 500"""
        )
        return {"history": [dict(r) for r in rows]}
    finally:
        await conn.close()


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    return JSONResponse({"detail": exc.detail}, status_code=exc.status_code)
